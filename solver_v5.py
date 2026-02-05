# solver_v5.py
"""
Solver de Distribuicao de PAX v5 - Modelo Duas Rodadas

Conceito:
- Rodada 1: Distribuir pax com origem TMIB
- Rodada 2: Distribuir pax com origem M9

Vantagens:
- Algoritmo mais simples (mesmo problema duas vezes)
- Capacidade total disponivel em cada rodada
- Prioridades faceis de implementar
"""

import json
import os
import re
import math
from dataclasses import dataclass, field
from typing import Dict, List, Set, Tuple, Optional
from copy import deepcopy
from itertools import permutations

from openpyxl import load_workbook


# ============================================================================
# CONFIGURACAO
# ============================================================================

DIST_FILE = "distplat.json"
SPEED_FILE = "velocidades.txt"
GANGWAY_FILE = "gangway.json"
INPUT_FILE = "solver_input.xlsx"
OUTPUT_FILE = "distribuicao_v5.txt"

DEFAULT_SPEED_KN = 14.0
MINUTES_PER_PAX = 1
MAX_DESVIO_NM = 3.0  # Desvio maximo aceitavel para parada no caminho

# Plataformas distantes (barco nao volta)
PLATAFORMAS_DISTANTES = {
    "PDO-01", "PDO-02", "PDO-03",
    "PGA-01", "PGA-02", "PGA-03", "PGA-04", "PGA-05", "PGA-06", "PGA-07", "PGA-08",
    "PRB-01"
}

# Clusters geograficos
CLUSTERS = {
    "M6_AREA": ["PCM-06", "PCM-08"],
    "B_CLUSTER": ["PCB-01", "PCB-02", "PCB-03", "PCB-04"],
    "M2M3": ["PCM-02", "PCM-03"],
    "M9_NEAR": ["PCM-04", "PCM-05", "PCM-09", "PCM-10", "PCM-11"],
    "M1M7": ["PCM-01", "PCM-07"],
    "PDO": ["PDO-01", "PDO-02", "PDO-03"],
    "PGA": ["PGA-01", "PGA-02", "PGA-03", "PGA-04", "PGA-05", "PGA-06", "PGA-07", "PGA-08"],
    "PRB": ["PRB-01"],
}

# Clusters que podem ser combinados em uma rota
# Baseado em CLAUDE.md: "todos perto de M9" podem ser combinados
CLUSTERS_COMPATIVEIS = [
    ("M6_AREA", "B_CLUSTER"),
    ("M6_AREA", "M9_NEAR"),
    ("M6_AREA", "M1M7"),
    ("M6_AREA", "M2M3"),
    ("B_CLUSTER", "M9_NEAR"),
    ("B_CLUSTER", "M1M7"),
    ("B_CLUSTER", "M2M3"),  # Adicionado: todos perto de M9
    ("M2M3", "M9_NEAR"),
    ("M2M3", "M1M7"),
    ("M1M7", "M9_NEAR"),
    ("PDO", "PGA"),
]

# Pares obrigatorios (devem ir juntos quando possivel)
PARES_OBRIGATORIOS = [
    ("PCM-02", "PCM-03"),
    ("PCM-06", "PCB-01"),
]


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class Demanda:
    plataforma: str
    plataforma_norm: str
    tmib: int = 0
    m9: int = 0
    prioridade: int = 99

    def total(self) -> int:
        return self.tmib + self.m9

    def copy(self):
        return Demanda(self.plataforma, self.plataforma_norm, self.tmib, self.m9, self.prioridade)


@dataclass
class Barco:
    nome: str
    disponivel: bool = False
    hora_saida: str = ""
    rota_fixa: str = ""
    velocidade: float = DEFAULT_SPEED_KN
    capacidade: int = 24

    def hora_saida_minutos(self) -> int:
        if not self.hora_saida or ":" not in self.hora_saida:
            return 999 * 60
        partes = self.hora_saida.split(":")
        return int(partes[0]) * 60 + int(partes[1])


@dataclass
class Config:
    troca_turma: bool = False
    rendidos_m9: int = 0


@dataclass
class ParadaExtra:
    plataforma: str
    qtd: int


@dataclass
class RotaDistante:
    barco: Barco
    destinos: List[Tuple[str, int, int]]  # (plataforma, tmib, m9)
    paradas_pre_m9: List[ParadaExtra] = field(default_factory=list)
    paradas_pos_m9: List[ParadaExtra] = field(default_factory=list)  # Entre M9 e destino final
    passa_por_m9: bool = False
    tmib_para_m9: int = 0
    m9_pickup: int = 0
    capacidade_extra: int = 0


@dataclass
class AlocacaoR1:
    barco: Barco
    destinos: List[Tuple[str, int]]  # (plataforma, qtd_tmib)
    tempo_chegada_m9: int = 0  # minutos desde meia-noite
    tmib_para_m9: int = 0  # pax TMIB que ficam em M9


@dataclass
class AlocacaoR2:
    barco: Barco
    destinos: List[Tuple[str, int]]  # (plataforma, qtd_m9)


# ============================================================================
# NORMALIZACAO DE PLATAFORMAS
# ============================================================================

def norm_plat(code: str) -> str:
    c = code.strip().upper()
    if c in ("TMIB", "NORWIND GALE"):
        return c
    if re.match(r"^(PCM|PCB|PGA|PRB|PDO)-\d{2}$", c):
        return c
    m = re.match(r"^M(\d+)$", c)
    if m:
        return f"PCM-{int(m.group(1)):02d}"
    b = re.match(r"^B(\d+)$", c)
    if b:
        return f"PCB-{int(b.group(1)):02d}"
    pg = re.match(r"^PGA(\d+)$", c)
    if pg:
        return f"PGA-{int(pg.group(1)):02d}"
    pdo = re.match(r"^PDO(\d+)$", c)
    if pdo:
        return f"PDO-{int(pdo.group(1)):02d}"
    prb = re.match(r"^PRB(\d+)$", c)
    if prb:
        return f"PRB-{int(prb.group(1)):02d}"
    return c


def short_plat(norm: str) -> str:
    n = norm.upper().strip()
    if n == "TMIB":
        return "TMIB"
    if n == "PCM-09":
        return "M9"
    if re.match(r"^PCM-\d{2}$", n):
        return f"M{int(n.split('-')[1])}"
    if re.match(r"^PCB-\d{2}$", n):
        return f"B{int(n.split('-')[1])}"
    if re.match(r"^PGA-\d{2}$", n):
        return f"PGA{int(n.split('-')[1])}"
    if re.match(r"^PDO-\d{2}$", n):
        return f"PDO{int(n.split('-')[1])}"
    if re.match(r"^PRB-\d{2}$", n):
        return f"PRB{int(n.split('-')[1])}"
    return n


def eh_plataforma_distante(plataforma_norm: str) -> bool:
    return plataforma_norm in PLATAFORMAS_DISTANTES


def obter_cluster(plataforma_norm: str) -> str:
    for nome_cluster, plataformas in CLUSTERS.items():
        if plataforma_norm in plataformas:
            return nome_cluster
    return "OTHER"


def clusters_compativeis(c1: str, c2: str) -> bool:
    if c1 == c2:
        return True
    for a, b in CLUSTERS_COMPATIVEIS:
        if (c1 == a and c2 == b) or (c1 == b and c2 == a):
            return True
    return False


# ============================================================================
# CARREGAMENTO DE DADOS
# ============================================================================

def load_distances(path: str) -> Dict[str, Dict[str, float]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    dist = {}
    for a, m in data.items():
        aN = norm_plat(str(a))
        dist.setdefault(aN, {})
        for b, v in m.items():
            dist[aN][norm_plat(str(b))] = float(v)
    return dist


def get_dist(distances: Dict, a_norm: str, b_norm: str) -> float:
    if a_norm == b_norm:
        return 0.0
    if a_norm in distances and b_norm in distances[a_norm]:
        return distances[a_norm][b_norm]
    if b_norm in distances and a_norm in distances[b_norm]:
        return distances[b_norm][a_norm]
    return 999.0


def load_speeds(path: str) -> Dict[str, float]:
    speeds = {}
    if not os.path.exists(path):
        return speeds
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    section = None
    for line in content.split("\n"):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        sm = re.match(r"\[([A-Z_]+)\]", line)
        if sm:
            section = sm.group(1).replace("_", " ")
            continue
        if "=" in line:
            parts = line.split("=")
            if len(parts) == 2:
                left = parts[0].strip()
                try:
                    spd = float(parts[1].strip())
                except ValueError:
                    continue
                if section and left.isdigit():
                    speeds[f"{section} {left}".upper()] = spd
                else:
                    name = left.replace("_", " ").upper()
                    speeds[name] = spd
                    speeds[left.upper()] = spd
    return speeds


def get_speed(speeds: Dict, name: str) -> float:
    up = name.upper()
    for v in [up, up.replace("_", " "), up.replace(" ", "_")]:
        if v in speeds:
            return speeds[v]
    return DEFAULT_SPEED_KN


def get_capacidade(name: str) -> int:
    up = name.upper()
    if "AQUA" in up and "HELIX" in up:
        return 100
    return 24


def eh_aqua_helix(name: str) -> bool:
    up = name.upper()
    return "AQUA" in up and "HELIX" in up


def load_gangway(path: str) -> Set[str]:
    if not os.path.exists(path):
        return set()
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    plataformas = data.get("plataformas_gangway", [])
    return {norm_plat(p) for p in plataformas}


def read_solver_input(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    config = Config()
    troca = ws.cell(row=4, column=3).value
    config.troca_turma = str(troca).strip().upper() == "SIM" if troca else False
    rend = ws.cell(row=5, column=3).value
    config.rendidos_m9 = int(rend) if rend else 0

    barcos = []
    r = 9
    while True:
        nome = ws.cell(row=r, column=2).value
        if not nome or str(nome).strip() == "":
            break
        disp = ws.cell(row=r, column=3).value
        hora = ws.cell(row=r, column=4).value
        rota = ws.cell(row=r, column=5).value

        if hora and hasattr(hora, 'strftime'):
            hora = hora.strftime("%H:%M")
        elif hora and ":" not in str(hora):
            try:
                h_val = float(hora) * 24
                hora = f"{int(h_val):02d}:{int((h_val % 1) * 60):02d}"
            except:
                hora = ""
        else:
            hora = str(hora).strip() if hora else ""

        rota_str = str(rota).strip() if rota and str(rota).strip().upper() != "NONE" else ""

        barcos.append(Barco(
            nome=str(nome).strip(),
            disponivel=str(disp).strip().upper() == "SIM" if disp else False,
            hora_saida=hora,
            rota_fixa=rota_str,
            capacidade=get_capacidade(str(nome).strip()),
        ))
        r += 1

    # Secao de demanda
    demand_header_row = r + 1
    demand_col_row = demand_header_row + 1
    demand_start = demand_col_row + 1

    demandas = []
    r = demand_start
    while True:
        plat = ws.cell(row=r, column=2).value
        if not plat or str(plat).strip() == "":
            break
        plat_str = str(plat).strip()
        m9_val = ws.cell(row=r, column=3).value
        tmib_val = ws.cell(row=r, column=4).value
        prio_val = ws.cell(row=r, column=5).value

        demandas.append(Demanda(
            plataforma=plat_str,
            plataforma_norm=norm_plat(plat_str),
            tmib=int(tmib_val) if tmib_val else 0,
            m9=int(m9_val) if m9_val else 0,
            prioridade=int(prio_val) if prio_val else 99,
        ))
        r += 1

    return config, barcos, demandas


# ============================================================================
# FUNCOES AUXILIARES
# ============================================================================

def tempo_viagem_minutos(distancia_nm: float, velocidade_kn: float) -> int:
    if velocidade_kn <= 0:
        return 999
    return math.ceil(distancia_nm / velocidade_kn * 60)


def calcular_desvio(plataforma_norm: str, distances: Dict) -> float:
    """Calcula desvio para parar em plataforma no caminho TMIB -> M9."""
    direto = get_dist(distances, "TMIB", norm_plat("M9"))
    via_plat = get_dist(distances, "TMIB", plataforma_norm) + get_dist(distances, plataforma_norm, norm_plat("M9"))
    return via_plat - direto


def calcular_melhor_posicao(plataforma_norm: str, destino_final_norm: str, distances: Dict) -> str:
    """
    Decide se plataforma fica melhor pré-M9 ou pós-M9 para rota distante.
    Retorna 'pre' ou 'pos'.
    """
    m9 = norm_plat("M9")

    # Desvio pré-M9: TMIB -> plat -> M9 vs TMIB -> M9
    desvio_pre = calcular_desvio(plataforma_norm, distances)

    # Desvio pós-M9: M9 -> plat -> destino vs M9 -> destino
    direto_pos = get_dist(distances, m9, destino_final_norm)
    via_plat_pos = get_dist(distances, m9, plataforma_norm) + get_dist(distances, plataforma_norm, destino_final_norm)
    desvio_pos = via_plat_pos - direto_pos

    # Escolher a posição com menor desvio
    if desvio_pre <= desvio_pos:
        return 'pre'
    else:
        return 'pos'


def escolher_destino_final(rota: RotaDistante, distances: Dict) -> str:
    """Escolhe um destino final representativo (mais distante de M9)."""
    m9 = norm_plat("M9")
    if not rota.destinos:
        return m9
    melhor = rota.destinos[0][0]
    melhor_dist = get_dist(distances, m9, melhor)
    for plat, _, _ in rota.destinos[1:]:
        dist = get_dist(distances, m9, plat)
        if dist > melhor_dist:
            melhor_dist = dist
            melhor = plat
    return melhor


def ordenar_por_tsp(destinos: List[Tuple[str, int]], origem: str, distances: Dict) -> List[Tuple[str, int]]:
    """Ordena destinos pelo vizinho mais proximo (TSP greedy)."""
    if len(destinos) <= 1:
        return list(destinos)

    # Para poucos destinos, testar todas as permutacoes
    if len(destinos) <= 6:
        melhor_ordem = None
        melhor_dist = float('inf')

        for perm in permutations(destinos):
            dist_total = 0.0
            atual = origem
            for plat, _ in perm:
                dist_total += get_dist(distances, atual, plat)
                atual = plat
            if dist_total < melhor_dist:
                melhor_dist = dist_total
                melhor_ordem = list(perm)

        return melhor_ordem

    # Para muitos destinos, usar vizinho mais proximo
    restantes = list(destinos)
    ordem = []
    atual = origem

    while restantes:
        mais_proximo = min(restantes, key=lambda d: get_dist(distances, atual, d[0]))
        ordem.append(mais_proximo)
        atual = mais_proximo[0]
        restantes.remove(mais_proximo)

    return ordem


def eh_compativel_com_rota(destinos_atuais: List[Tuple[str, int]], novo_destino_norm: str) -> bool:
    """Verifica se novo destino e compativel com clusters ja na rota."""
    if not destinos_atuais:
        return True

    cluster_novo = obter_cluster(novo_destino_norm)

    for plat, _ in destinos_atuais:
        cluster_atual = obter_cluster(plat)
        if not clusters_compativeis(cluster_atual, cluster_novo):
            return False

    return True


# ============================================================================
# FASE 0: PREPARACAO
# ============================================================================

def preparar_dados(demandas: List[Demanda]) -> Tuple[int, List[Demanda], List[Demanda]]:
    """Separa demanda em M9, proximas e distantes."""
    demanda_m9_para_m9 = 0
    demanda_proximas = []
    demanda_distantes = []

    for d in demandas:
        if short_plat(d.plataforma_norm) == "M9":
            demanda_m9_para_m9 = d.tmib
        elif eh_plataforma_distante(d.plataforma_norm):
            if d.tmib > 0 or d.m9 > 0:
                demanda_distantes.append(d.copy())
        else:
            if d.tmib > 0 or d.m9 > 0:
                demanda_proximas.append(d.copy())

    return demanda_m9_para_m9, demanda_proximas, demanda_distantes


# ============================================================================
# FASE 1: ALOCAR ROTAS DISTANTES
# ============================================================================

def agrupar_por_cluster_distante(demandas: List[Demanda]) -> Dict[str, List[Demanda]]:
    """Agrupa demandas distantes por cluster."""
    grupos = {}
    for d in demandas:
        cluster = obter_cluster(d.plataforma_norm)
        if cluster not in grupos:
            grupos[cluster] = []
        grupos[cluster].append(d)
    return grupos


def alocar_rotas_distantes(
    demanda_distantes: List[Demanda],
    barcos: List[Barco],
    distances: Dict
) -> Tuple[List[RotaDistante], List[Barco]]:
    """Aloca barcos para clusters distantes."""

    rotas = []
    barcos_usados = []

    if not demanda_distantes:
        return rotas, barcos_usados

    clusters = agrupar_por_cluster_distante(demanda_distantes)

    # Ordenar barcos por hora de saida (ultimo primeiro para rotas distantes)
    barcos_ordenados = sorted(barcos, key=lambda b: -b.hora_saida_minutos())

    for nome_cluster, demandas_cluster in clusters.items():
        total_pax = sum(d.tmib + d.m9 for d in demandas_cluster)
        if total_pax == 0:
            continue

        # Selecionar barco disponivel
        barco = None
        for b in barcos_ordenados:
            if b not in barcos_usados:
                barco = b
                break

        if barco is None:
            print(f"  AVISO: Sem barco para cluster distante {nome_cluster}")
            continue

        # Verificar se tem demanda M9 (precisa passar por M9)
        tem_demanda_m9 = any(d.m9 > 0 for d in demandas_cluster)
        # Para clusters distantes (PDO/PGA/PRB), sempre passar por M9
        if nome_cluster in ["PDO", "PGA", "PRB"]:
            tem_demanda_m9 = True
        total_tmib = sum(d.tmib for d in demandas_cluster)
        total_m9 = sum(d.m9 for d in demandas_cluster)

        # Capacidade extra = vagas em TMIB (pax M9 embarcam em M9, nao competem)
        capacidade_extra = barco.capacidade - total_tmib

        destinos = [(d.plataforma_norm, d.tmib, d.m9) for d in demandas_cluster]

        rota = RotaDistante(
            barco=barco,
            destinos=destinos,
            passa_por_m9=tem_demanda_m9,
            tmib_para_m9=0,
            m9_pickup=total_m9,
            capacidade_extra=max(0, capacidade_extra),
        )

        rotas.append(rota)
        barcos_usados.append(barco)
        print(f"  Rota distante {barco.nome}: {nome_cluster} ({total_pax} pax)")

    return rotas, barcos_usados


# ============================================================================
# FASE 1.5: BALANCEAR CAPACIDADE
# ============================================================================

def balancear_capacidade(
    demanda_proximas: List[Demanda],
    rotas_distantes: List[RotaDistante],
    barcos_livres: List[Barco],
    distances: Dict,
    demanda_m9_para_m9: int = 0
) -> Tuple[List[Demanda], int]:
    """Resolve gargalos usando capacidade extra das rotas distantes.
    Retorna (demandas_atualizadas, pax_m9_alocados_em_distantes)."""

    capacidade_total = sum(b.capacidade for b in barcos_livres)
    demanda_total_tmib = sum(d.tmib for d in demanda_proximas) + demanda_m9_para_m9
    pax_m9_em_distantes = 0

    if demanda_total_tmib <= capacidade_total:
        return demanda_proximas, 0  # Sem gargalo

    excedente = demanda_total_tmib - capacidade_total
    print(f"  Gargalo detectado: {excedente} pax excedentes (inclui {demanda_m9_para_m9} para M9)")

    # Ordenar plataformas por desvio (menor primeiro)
    candidatas = sorted(
        [d for d in demanda_proximas if d.tmib > 0],
        key=lambda d: calcular_desvio(d.plataforma_norm, distances)
    )

    for rota in rotas_distantes:
        if rota.capacidade_extra <= 0 or excedente <= 0:
            continue

        for d in candidatas:
            if d.tmib <= 0 or excedente <= 0:
                continue

            desvio = calcular_desvio(d.plataforma_norm, distances)
            if desvio > MAX_DESVIO_NM:
                continue

            pax_mover = min(d.tmib, rota.capacidade_extra, excedente)

            if pax_mover > 0:
                # Decidir se parada fica melhor pré ou pós M9
                destino_final = escolher_destino_final(rota, distances)
                cluster = obter_cluster(d.plataforma_norm)
                rota_e_distante = any(eh_plataforma_distante(plat) for plat, _, _ in rota.destinos)
                if rota.passa_por_m9 and rota_e_distante:
                    # Em rotas distantes, evitar pre-M9 para garantir M9 antes
                    posicao = 'pos'
                elif cluster == "M6_AREA" and rota.passa_por_m9:
                    posicao = 'pos'
                else:
                    # Para plataformas perto de M9, preferir pré-M9 se o desvio não for pior
                    if cluster == "M9_NEAR" and rota.passa_por_m9:
                        desvio_pre = calcular_desvio(d.plataforma_norm, distances)
                        m9 = norm_plat("M9")
                        direto_pos = get_dist(distances, m9, destino_final)
                        via_pos = get_dist(distances, m9, d.plataforma_norm) + get_dist(distances, d.plataforma_norm, destino_final)
                        desvio_pos = via_pos - direto_pos
                        if desvio_pre <= MAX_DESVIO_NM and desvio_pre <= desvio_pos + 0.5:
                            posicao = 'pre'
                        else:
                            posicao = calcular_melhor_posicao(d.plataforma_norm, destino_final, distances)
                    else:
                        posicao = calcular_melhor_posicao(d.plataforma_norm, destino_final, distances)

                if posicao == 'pre':
                    rota.paradas_pre_m9.append(ParadaExtra(d.plataforma_norm, pax_mover))
                    print(f"    Movido {pax_mover} pax de {short_plat(d.plataforma_norm)} para rota {rota.barco.nome} (pré-M9)")
                else:
                    rota.paradas_pos_m9.append(ParadaExtra(d.plataforma_norm, pax_mover))
                    print(f"    Movido {pax_mover} pax de {short_plat(d.plataforma_norm)} para rota {rota.barco.nome} (pós-M9)")

                d.tmib -= pax_mover
                rota.capacidade_extra -= pax_mover
                excedente -= pax_mover

    # Se ainda há excedente e pax M9 para alocar, colocar nas rotas distantes
    if excedente > 0 and demanda_m9_para_m9 > 0:
        for rota in rotas_distantes:
            if rota.capacidade_extra <= 0 or excedente <= 0:
                continue
            # Alocar pax TMIB→M9 na rota distante (que já passa por M9)
            pax_mover = min(demanda_m9_para_m9 - pax_m9_em_distantes, rota.capacidade_extra, excedente)
            if pax_mover > 0 and rota.passa_por_m9:
                rota.tmib_para_m9 += pax_mover
                rota.capacidade_extra -= pax_mover
                pax_m9_em_distantes += pax_mover
                excedente -= pax_mover
                print(f"    Movido {pax_mover} pax TMIB→M9 para rota {rota.barco.nome}")

    if excedente > 0:
        print(f"  AVISO: {excedente} pax TMIB ainda sem alocacao")

    return demanda_proximas, pax_m9_em_distantes


# ============================================================================
# FASE 2: ALOCAR RODADA 1 (TMIB -> PROXIMAS)
# ============================================================================

def formar_pacotes(demandas: List[Demanda], capacidade_max: int = 24) -> List[List[Demanda]]:
    """Forma pacotes respeitando pares obrigatorios (divide se exceder capacidade)."""
    pacotes = []
    usados = set()

    # Pares obrigatorios primeiro
    for p1, p2 in PARES_OBRIGATORIOS:
        d1 = next((d for d in demandas if d.plataforma_norm == p1 and d.tmib > 0), None)
        d2 = next((d for d in demandas if d.plataforma_norm == p2 and d.tmib > 0), None)

        if d1 and d2:
            total_par = d1.tmib + d2.tmib
            if total_par <= capacidade_max:
                # Par cabe em um barco
                pacotes.append([d1, d2])
                usados.add(p1)
                usados.add(p2)
            else:
                # Par NAO cabe - tratar como individuais
                print(f"  AVISO: Par {short_plat(p1)}+{short_plat(p2)} ({total_par} pax) excede capacidade, dividindo")
                # Nao adiciona em usados, serao tratados como individuais

    # Demandas individuais
    for d in demandas:
        if d.plataforma_norm not in usados and d.tmib > 0:
            pacotes.append([d])

    return pacotes


def alocar_rodada1(
    demanda_proximas: List[Demanda],
    barcos_livres: List[Barco],
    distances: Dict,
    demanda_m9_para_m9: int = 0
) -> List[AlocacaoR1]:
    """Aloca demanda TMIB nos barcos (bin-packing)."""

    alocacoes = []
    pacotes = formar_pacotes(demanda_proximas)

    print(f"  Pacotes formados: {len(pacotes)}")
    for i, p in enumerate(pacotes):
        plats = "+".join(short_plat(d.plataforma_norm) for d in p)
        total = sum(d.tmib for d in p)
        print(f"    [{i}] {plats}: {total} pax")

    # Ordenar pacotes por tamanho (maior primeiro)
    pacotes.sort(key=lambda p: -sum(d.tmib for d in p))

    # Ordenar barcos por hora de saida
    barcos_ordenados = sorted(barcos_livres, key=lambda b: b.hora_saida_minutos())

    # Rastrear pacotes alocados por indice
    pacotes_alocados = set()

    for barco in barcos_ordenados:
        print(f"  Processando barco {barco.nome} (cap={barco.capacidade}):")
        destinos = []
        capacidade_restante = barco.capacidade

        for idx, pacote in enumerate(pacotes):
            if idx in pacotes_alocados:
                continue

            total_pacote = sum(d.tmib for d in pacote)
            if total_pacote > capacidade_restante:
                continue

            # Verificar compatibilidade de cluster
            compativel = True
            for d in pacote:
                if not eh_compativel_com_rota(destinos, d.plataforma_norm):
                    compativel = False
                    break

            if compativel:
                for d in pacote:
                    if d.tmib > 0:
                        destinos.append((d.plataforma_norm, d.tmib))
                capacidade_restante -= total_pacote
                pacotes_alocados.add(idx)
                plats = "+".join(short_plat(d.plataforma_norm) for d in pacote)
                print(f"    + Alocado pacote {plats} ({total_pacote} pax), restam {capacidade_restante}")

        if destinos:
            alocacoes.append(AlocacaoR1(barco=barco, destinos=destinos))
        else:
            print(f"    Nenhum pacote alocado para {barco.nome}!")

    # Verificar pacotes nao alocados
    for idx, pacote in enumerate(pacotes):
        if idx not in pacotes_alocados:
            plats = ", ".join(short_plat(d.plataforma_norm) for d in pacote)
            total = sum(d.tmib for d in pacote)
            if total > 0:
                print(f"  AVISO: Pacote nao alocado: {plats} ({total} pax)")

    # Distribuir pax TMIB→M9 entre os barcos com espaço
    if demanda_m9_para_m9 > 0:
        restante_m9 = demanda_m9_para_m9
        for aloc in alocacoes:
            if restante_m9 <= 0:
                break
            carga_atual = sum(q for _, q in aloc.destinos)
            espaco = aloc.barco.capacidade - carga_atual
            if espaco > 0:
                pax_m9 = min(espaco, restante_m9)
                aloc.tmib_para_m9 = pax_m9
                restante_m9 -= pax_m9
                print(f"    + {aloc.barco.nome} leva {pax_m9} pax TMIB→M9")

        if restante_m9 > 0:
            print(f"  AVISO: {restante_m9} pax TMIB→M9 nao alocados")

    return alocacoes


# ============================================================================
# FASE 2.5: SIMULAR RODADA 1 E CALCULAR CHEGADA EM M9
# ============================================================================

def simular_rodada1(
    alocacoes: List[AlocacaoR1],
    distances: Dict
) -> List[AlocacaoR1]:
    """Simula R1, ordena paradas por TSP e calcula tempo de chegada em M9."""

    m9_norm = norm_plat("M9")

    for aloc in alocacoes:
        if not aloc.destinos:
            continue

        # Ordenar destinos por TSP
        aloc.destinos = ordenar_por_tsp(aloc.destinos, "TMIB", distances)

        # Calcular tempo
        tempo_atual = aloc.barco.hora_saida_minutos()
        posicao_atual = "TMIB"

        for plat, qtd in aloc.destinos:
            dist = get_dist(distances, posicao_atual, plat)
            tempo_atual += tempo_viagem_minutos(dist, aloc.barco.velocidade)
            tempo_atual += qtd * MINUTES_PER_PAX
            posicao_atual = plat

        # Tempo para chegar em M9
        dist_para_m9 = get_dist(distances, posicao_atual, m9_norm)
        aloc.tempo_chegada_m9 = tempo_atual + tempo_viagem_minutos(dist_para_m9, aloc.barco.velocidade)

    return alocacoes


# ============================================================================
# FASE 3: ALOCAR RODADA 2 (M9 -> PROXIMAS)
# ============================================================================

def alocar_rodada2(
    demanda_proximas: List[Demanda],
    alocacoes_r1: List[AlocacaoR1],
    distances: Dict
) -> List[AlocacaoR2]:
    """Aloca demanda M9 nos barcos por ordem de chegada em M9."""

    alocacoes = []

    # Filtrar demanda M9 de proximas
    demandas_m9 = [d.copy() for d in demanda_proximas if d.m9 > 0]

    if not demandas_m9:
        return alocacoes

    # Ordenar barcos por tempo de chegada em M9
    barcos_ordenados = sorted(alocacoes_r1, key=lambda a: a.tempo_chegada_m9)

    # Ordenar demanda por prioridade
    demandas_m9.sort(key=lambda d: d.prioridade)

    for aloc_r1 in barcos_ordenados:
        destinos = []
        capacidade_restante = aloc_r1.barco.capacidade

        for d in demandas_m9:
            if d.m9 <= 0:
                continue

            if d.m9 <= capacidade_restante:
                if eh_compativel_com_rota(destinos, d.plataforma_norm):
                    destinos.append((d.plataforma_norm, d.m9))
                    capacidade_restante -= d.m9
                    d.m9 = 0

        if destinos:
            alocacoes.append(AlocacaoR2(barco=aloc_r1.barco, destinos=destinos))

    # Verificar demanda M9 nao alocada
    for d in demandas_m9:
        if d.m9 > 0:
            print(f"  AVISO: Demanda M9 nao alocada: {short_plat(d.plataforma_norm)} = {d.m9}")

    return alocacoes


# ============================================================================
# FASE 4: GERAR ROTAS FINAIS
# ============================================================================

def gerar_string_rota_duas_rodadas(
    aloc_r1: AlocacaoR1,
    aloc_r2: Optional[AlocacaoR2],
    distances: Dict
) -> str:
    """Gera string de rota para barcos que fazem duas rodadas."""

    partes = []
    m9_norm = norm_plat("M9")

    # TMIB com total de pax (inclui pax que ficam em M9)
    total_tmib = sum(qtd for _, qtd in aloc_r1.destinos) + aloc_r1.tmib_para_m9
    partes.append(f"TMIB +{total_tmib}")

    # Paradas R1
    for plat, qtd in aloc_r1.destinos:
        partes.append(f"{short_plat(plat)} -{qtd}")

    # M9 (entrega pax TMIB→M9 e/ou pega pax M9)
    tem_r2 = aloc_r2 and aloc_r2.destinos
    if aloc_r1.tmib_para_m9 > 0 or tem_r2:
        m9_str = "M9"
        if aloc_r1.tmib_para_m9 > 0:
            m9_str += f" -{aloc_r1.tmib_para_m9}"
        if tem_r2:
            destinos_r2 = ordenar_por_tsp(aloc_r2.destinos, m9_norm, distances)
            total_m9 = sum(qtd for _, qtd in destinos_r2)
            m9_str += f" +{total_m9}"
        partes.append(m9_str)

        # Paradas R2
        if tem_r2:
            for plat, qtd in destinos_r2:
                partes.append(f"{short_plat(plat)} (-{qtd})")

    return "/".join(partes)


def gerar_string_rota_distante(rota: RotaDistante, distances: Dict) -> str:
    """Gera string de rota para destinos distantes."""

    partes = []
    m9_norm = norm_plat("M9")

    # Calcular total de pax TMIB (pax M9 embarcam em M9, nao aqui)
    total_pax = sum(t for _, t, m in rota.destinos)  # So TMIB dos destinos finais
    total_pax += sum(p.qtd for p in rota.paradas_pre_m9)
    total_pax += sum(p.qtd for p in rota.paradas_pos_m9)
    total_pax += rota.tmib_para_m9  # Pax TMIB que ficam em M9

    partes.append(f"TMIB +{total_pax}")

    # Paradas pre-M9 (ordenar por TSP a partir de TMIB)
    if rota.paradas_pre_m9:
        paradas_pre = ordenar_por_tsp(
            [(p.plataforma, p.qtd) for p in rota.paradas_pre_m9],
            "TMIB",
            distances
        )
        for plat, qtd in paradas_pre:
            partes.append(f"{short_plat(plat)} -{qtd}")

    # M9 (se passa ou tem pax para entregar)
    if rota.passa_por_m9 or rota.tmib_para_m9 > 0 or rota.paradas_pos_m9:
        m9_str = "M9"
        if rota.tmib_para_m9 > 0:
            m9_str += f" -{rota.tmib_para_m9}"
        if rota.m9_pickup > 0:
            m9_str += f" +{rota.m9_pickup}"
        partes.append(m9_str)

    # Paradas pos-M9 + destinos finais (ordenar juntos por TSP a partir de M9)
    destinos_pos_m9 = []

    # Adicionar paradas pos-M9
    for p in rota.paradas_pos_m9:
        destinos_pos_m9.append((p.plataforma, p.qtd, 0, 'extra'))  # (plat, tmib, m9, tipo)

    # Adicionar destinos finais
    for plat, tmib, m9 in rota.destinos:
        destinos_pos_m9.append((plat, tmib, m9, 'final'))

    # Ordenar por TSP a partir de M9
    if destinos_pos_m9:
        destinos_ordenados = ordenar_por_tsp(
            [(d[0], d[1] + d[2]) for d in destinos_pos_m9],
            m9_norm,
            distances
        )

        # Mapear de volta para dados completos
        destinos_map = {d[0]: (d[1], d[2], d[3]) for d in destinos_pos_m9}

        for plat, _ in destinos_ordenados:
            tmib, m9, tipo = destinos_map[plat]
            dest_str = short_plat(plat)
            if tmib > 0:
                dest_str += f" -{tmib}"
            if m9 > 0:
                dest_str += f" (-{m9})"
            partes.append(dest_str)

    return "/".join(partes)


def gerar_rotas_finais(
    alocacoes_r1: List[AlocacaoR1],
    alocacoes_r2: List[AlocacaoR2],
    rotas_distantes: List[RotaDistante],
    distances: Dict
) -> List[Tuple[Barco, str]]:
    """Gera todas as strings de rota."""

    resultados = []

    # Mapear R2 por barco
    r2_por_barco = {a.barco.nome: a for a in alocacoes_r2}

    # Rotas de duas rodadas
    for aloc_r1 in alocacoes_r1:
        aloc_r2 = r2_por_barco.get(aloc_r1.barco.nome)
        rota_str = gerar_string_rota_duas_rodadas(aloc_r1, aloc_r2, distances)
        resultados.append((aloc_r1.barco, rota_str))

    # Rotas distantes
    for rota in rotas_distantes:
        rota_str = gerar_string_rota_distante(rota, distances)
        resultados.append((rota.barco, rota_str))

    # Ordenar por hora de saida
    resultados.sort(key=lambda x: x[0].hora_saida_minutos())

    return resultados


# ============================================================================
# VALIDACAO
# ============================================================================

def validar_resultado(
    resultados: List[Tuple[Barco, str]],
    demandas: List[Demanda]
) -> List[str]:
    """Valida o resultado e retorna avisos."""

    avisos = []

    # Calcular entregas por plataforma
    entregas_tmib = {}
    entregas_m9 = {}

    for barco, rota_str in resultados:
        partes = rota_str.split("/")
        for parte in partes:
            parte = parte.strip()
            tokens = parte.split()
            if not tokens:
                continue

            plat = tokens[0]
            if plat == "TMIB":
                continue

            plat_norm = norm_plat(plat)

            for token in tokens[1:]:
                # TMIB drop: -N
                m = re.match(r'^-(\d+)$', token)
                if m:
                    qtd = int(m.group(1))
                    entregas_tmib[plat_norm] = entregas_tmib.get(plat_norm, 0) + qtd

                # M9 drop: (-N)
                m = re.match(r'^\(-(\d+)\)$', token)
                if m:
                    qtd = int(m.group(1))
                    entregas_m9[plat_norm] = entregas_m9.get(plat_norm, 0) + qtd

    # Comparar com demanda
    for d in demandas:
        if short_plat(d.plataforma_norm) == "M9":
            continue

        entregue_tmib = entregas_tmib.get(d.plataforma_norm, 0)
        entregue_m9 = entregas_m9.get(d.plataforma_norm, 0)

        if entregue_tmib < d.tmib:
            avisos.append(f"Faltam {d.tmib - entregue_tmib} pax TMIB para {short_plat(d.plataforma_norm)}")

        if entregue_m9 < d.m9:
            avisos.append(f"Faltam {d.m9 - entregue_m9} pax M9 para {short_plat(d.plataforma_norm)}")

    return avisos


# ============================================================================
# FUNCAO PRINCIPAL
# ============================================================================

def resolver_distribuicao(
    config: Config,
    barcos: List[Barco],
    demandas: List[Demanda],
    distances: Dict
) -> Tuple[List[Tuple[Barco, str]], List[str]]:
    """Resolve a distribuicao de pax usando o modelo de duas rodadas."""

    avisos = []

    # FASE 0: Preparar dados
    print("\nFASE 0: Preparando dados...")
    demanda_m9_para_m9, demanda_proximas, demanda_distantes = preparar_dados(demandas)

    barcos_disponiveis = [b for b in barcos if b.disponivel]
    barcos_disponiveis.sort(key=lambda b: b.hora_saida_minutos())

    print(f"  Barcos disponiveis: {len(barcos_disponiveis)}")
    print(f"  Demanda proximas: {sum(d.tmib for d in demanda_proximas)} TMIB + {sum(d.m9 for d in demanda_proximas)} M9")
    print(f"  Demanda distantes: {sum(d.tmib for d in demanda_distantes)} TMIB + {sum(d.m9 for d in demanda_distantes)} M9")

    # FASE 1: Alocar rotas distantes
    print("\nFASE 1: Alocando rotas distantes...")
    rotas_distantes, barcos_usados = alocar_rotas_distantes(
        demanda_distantes, barcos_disponiveis, distances
    )

    barcos_livres = [b for b in barcos_disponiveis if b not in barcos_usados]
    print(f"  Barcos livres para proximas: {len(barcos_livres)}")
    for b in barcos_livres:
        print(f"    - {b.nome} ({b.hora_saida})")

    # FASE 1.5: Balancear capacidade
    print("\nFASE 1.5: Balanceando capacidade...")
    demanda_proximas, pax_m9_em_distantes = balancear_capacidade(
        demanda_proximas, rotas_distantes, barcos_livres, distances, demanda_m9_para_m9
    )
    # Ajustar demanda M9 restante para alocacao na R1
    demanda_m9_para_m9_restante = demanda_m9_para_m9 - pax_m9_em_distantes

    # FASE 2: Alocar rodada 1
    print("\nFASE 2: Alocando Rodada 1 (TMIB)...")
    print(f"  Demanda TMIB restante por plataforma:")
    for d in demanda_proximas:
        if d.tmib > 0:
            print(f"    - {short_plat(d.plataforma_norm)}: {d.tmib}")
    if demanda_m9_para_m9_restante > 0:
        print(f"    - M9 (destino final): {demanda_m9_para_m9_restante}")
    alocacoes_r1 = alocar_rodada1(demanda_proximas, barcos_livres, distances, demanda_m9_para_m9_restante)
    print(f"  Barcos com alocacao R1: {len(alocacoes_r1)}")
    for aloc in alocacoes_r1:
        total = sum(q for _, q in aloc.destinos)
        dests = ", ".join(f"{short_plat(p)}:{q}" for p, q in aloc.destinos)
        print(f"    - {aloc.barco.nome}: {total} pax [{dests}]")

    # FASE 2.5: Simular rodada 1
    print("\nFASE 2.5: Simulando Rodada 1...")
    alocacoes_r1 = simular_rodada1(alocacoes_r1, distances)

    for aloc in alocacoes_r1:
        h = aloc.tempo_chegada_m9 // 60
        m = aloc.tempo_chegada_m9 % 60
        print(f"  {aloc.barco.nome}: chega em M9 as {h:02d}:{m:02d}")

    # FASE 3: Alocar rodada 2
    print("\nFASE 3: Alocando Rodada 2 (M9)...")
    alocacoes_r2 = alocar_rodada2(demanda_proximas, alocacoes_r1, distances)
    print(f"  Barcos com alocacao R2: {len(alocacoes_r2)}")

    # FASE 4: Gerar rotas finais
    print("\nFASE 4: Gerando rotas finais...")
    resultados = gerar_rotas_finais(alocacoes_r1, alocacoes_r2, rotas_distantes, distances)

    # Validacao
    avisos = validar_resultado(resultados, demandas)

    return resultados, avisos


# ============================================================================
# OUTPUT
# ============================================================================

def write_output(resultados: List[Tuple[Barco, str]], avisos: List[str], config: Config, path: str):
    """Escreve resultado no arquivo de saida."""

    lines = []
    for barco, rota in resultados:
        line = f"{barco.nome}  {barco.hora_saida}  {rota}"
        lines.append(line)

    # Calcular totais
    total_tmib = 0
    total_m9 = 0
    for _, rota in resultados:
        partes = rota.split("/")
        for parte in partes:
            for token in parte.split():
                m = re.match(r'^-(\d+)$', token)
                if m:
                    total_tmib += int(m.group(1))
                m = re.match(r'^\(-(\d+)\)$', token)
                if m:
                    total_m9 += int(m.group(1))

    with open(path, "w", encoding="utf-8") as f:
        f.write("DISTRIBUICAO DE PAX (Solver v5 - Duas Rodadas)\n")
        f.write("=" * 70 + "\n")
        if config.troca_turma:
            f.write(f"Troca de turma: SIM | Rendidos em M9: {config.rendidos_m9}\n")
        f.write("\n")

        for line in lines:
            f.write(line + "\n")

        f.write("\n" + "-" * 70 + "\n")
        f.write(f"Resumo: {total_tmib} pax TMIB + {total_m9} pax M9 = {total_tmib + total_m9} pax total\n")
        f.write(f"Barcos utilizados: {len(resultados)}\n")
        f.write("=" * 70 + "\n")

        if avisos:
            f.write("\nAVISOS:\n")
            for aviso in avisos:
                f.write(f"  - {aviso}\n")

    print(f"\nResultado salvo em: {path}\n")
    for line in lines:
        print(f"  {line}")

    print(f"\n  Resumo: {total_tmib} TMIB + {total_m9} M9 = {total_tmib + total_m9} pax | {len(resultados)} barcos")

    if avisos:
        print("\n  AVISOS:")
        for aviso in avisos:
            print(f"    - {aviso}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 70)
    print("SOLVER DE DISTRIBUICAO DE PAX v5 - Modelo Duas Rodadas")
    print("=" * 70)

    if not os.path.exists(INPUT_FILE):
        print(f"ERRO: '{INPUT_FILE}' nao encontrado.")
        return

    if not os.path.exists(DIST_FILE):
        print(f"ERRO: '{DIST_FILE}' nao encontrado.")
        return

    # Carregar dados
    distances = load_distances(DIST_FILE)
    speeds = load_speeds(SPEED_FILE)
    config, barcos, demandas = read_solver_input(INPUT_FILE)

    # Aplicar velocidades
    for barco in barcos:
        barco.velocidade = get_speed(speeds, barco.nome)

    # Info
    n_disponiveis = sum(1 for b in barcos if b.disponivel)
    total_tmib = sum(d.tmib for d in demandas)
    total_m9 = sum(d.m9 for d in demandas if short_plat(d.plataforma_norm) != "M9")

    print(f"\nConfiguracao:")
    print(f"  Troca de turma: {'SIM' if config.troca_turma else 'NAO'}")
    print(f"  Barcos disponiveis: {n_disponiveis}")
    print(f"  Demanda total: {total_tmib} TMIB + {total_m9} M9 = {total_tmib + total_m9} pax")

    if n_disponiveis == 0:
        print("\nERRO: Nenhum barco disponivel.")
        return

    # Resolver
    resultados, avisos = resolver_distribuicao(config, barcos, demandas, distances)

    if resultados:
        write_output(resultados, avisos, config, OUTPUT_FILE)


if __name__ == "__main__":
    main()
