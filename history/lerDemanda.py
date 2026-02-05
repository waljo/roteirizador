import openpyxl
import sys
import os

def _ensure_demanda_file(caminho: str):
    pasta = os.path.dirname(caminho)
    if pasta and not os.path.exists(pasta):
        os.makedirs(pasta, exist_ok=True)
    if not os.path.exists(caminho):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Plataforma", "TMIB", "PCM-09"])
        wb.save(caminho)
        print(f"Arquivo criado: {caminho}")


def ler_demanda(caminho="Demandas/demandas.xlsx"):
    _ensure_demanda_file(caminho)

    wb = openpyxl.load_workbook(caminho)
    ws = wb.active

    # Linha 1: cabeçalho (Plataforma | Origem1 | Origem2 ...)
    cabecalho = [cell.value for cell in ws[1]]
    origens = cabecalho[1:]  # nomes das origens (ex: PCM-09, TMIB)

    demandas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        plataforma = row[0]
        if not plataforma:
            continue
        pax_por_origem = {}
        for i, origem in enumerate(origens):
            valor = row[i + 1] if i + 1 < len(row) else 0
            pax_por_origem[origem] = valor if valor else 0
        demandas.append({
            "plataforma": plataforma,
            "pax": pax_por_origem
        })

    return origens, demandas


def exibir_demanda(origens, demandas):
    if not demandas:
        print("Nenhuma demanda encontrada.")
        return
    # Cabeçalho
    col_plat = max(len(d["plataforma"]) for d in demandas)
    col_plat = max(col_plat, len("Plataforma"))

    header = f"{'Plataforma':<{col_plat}}"
    for orig in origens:
        header += f"  {orig:>8}"
    header += f"  {'Total':>6}"

    print("=" * len(header))
    print("DEMANDA DE DISTRIBUIÇÃO DE PAX")
    print("=" * len(header))
    print(header)
    print("-" * len(header))

    total_geral = {orig: 0 for orig in origens}

    for d in demandas:
        linha = f"{d['plataforma']:<{col_plat}}"
        total_plat = 0
        for orig in origens:
            val = d["pax"][orig]
            linha += f"  {val:>8}"
            total_plat += val
            total_geral[orig] += val
        linha += f"  {total_plat:>6}"
        print(linha)

    print("-" * len(header))
    totais = f"{'TOTAL':<{col_plat}}"
    soma = 0
    for orig in origens:
        totais += f"  {total_geral[orig]:>8}"
        soma += total_geral[orig]
    totais += f"  {soma:>6}"
    print(totais)
    print("=" * len(header))


if __name__ == "__main__":
    caminho = sys.argv[1] if len(sys.argv) > 1 else "Demandas/demandas.xlsx"
    origens, demandas = ler_demanda(caminho)
    exibir_demanda(origens, demandas)
