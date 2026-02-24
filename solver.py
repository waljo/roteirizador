# solver.py
"""
Solver de Distribuição Automática de PAX v4
- Pares obrigatórios (M2+M3, M6+B1) mantidos como unidade atômica
- Otimização combinatória: testa todas as atribuições pacote→barco
- AQUA Helix priorizada para rotas diretas de alta capacidade
- Clustering geográfico e nearest-neighbor para ordenação de paradas
"""

import json
import os
import re
import math
from dataclasses import dataclass, field
from typing import Dict, List, Set, Tuple, Optional
from copy import deepcopy

from openpyxl import load_workbook


DIST_FILE = "distplat.json"
SPEED_FILE = "velocidades.txt"
GANGWAY_FILE = "gangway.json"
INPUT_FILE = "solver_input.xlsx"
OUTPUT_FILE = "distribuicao.txt"

DEFAULT_SPEED_KN = 14.0
AQUA_APPROACH_TIME = 25  # minutos por parada
MINUTES_PER_PAX = 1  # minuto por pax embarcado/desembarcado
M9_CONSOLIDATION_PENALTY_NM = 5.0  # penalidade por espalhar embarques M9 em varios barcos
ENABLE_DISTANT_CLUSTER_DEDICATION = False  # evita reservar barco e reduzir capacidade total
PRIORITY_TIME_WEIGHT = 0.05  # peso (NM-equivalente por minuto) para antecipar prioridades
COMFORT_PAX_MIN_WEIGHT = 0.02  # peso (NM-equivalente por pax-minuto) para conforto
PAX_ARRIVAL_WEIGHT = 0.1  # peso forte (NM-equivalente por pax-minuto) para priorizar grandes entregas cedo
BACKTRACK_PENALTY_NM = 10.0  # penalidade por voltar em direção ao início (evitar "ir longe e voltar")
PRIORITY1_PRECEDENCE_PENALTY_NM = 250.0  # penalidade alta para colocar prioridade 1 depois de nao-prioridade
PRIORITY1_PRE_M9_MAX_DETOUR_NM = 1.5  # promove prioridade 1 para pre-M9 apenas com desvio pequeno
PRIORITY_MIX_FIT_PENALTY_NM = 120.0  # penaliza separar P2/P3 de barco com P1 quando caberia junto
CLUSTER_SWITCH_PENALTY_NM = 8.0  # penaliza mudanca de cluster na mesma perna
INCOMPATIBLE_CLUSTER_SWITCH_PENALTY_NM = 24.0  # salto entre clusters incompativeis
CROSS_CLUSTER_JUMP_PENALTY_PER_NM = 4.0  # penalidade por NM de salto entre clusters
CROSS_CLUSTER_JUMP_FREE_NM = 1.5  # folga sem penalidade para transicoes pequenas

# Clusters geográficos baseados em proximidade real
GEO_CLUSTERS = {
    "M6_AREA": ["PCM-06", "PCM-08"],  # M6/M8 muito próximos (0.42 NM)
    "B_CLUSTER": ["PCB-01", "PCB-02", "PCB-03", "PCB-04"],  # B1-B4 próximos
    "M2M3": ["PCM-02", "PCM-03"],  # M2/M3 próximos (1.04 NM)
    "M9_NEAR": ["PCM-04", "PCM-05", "PCM-09", "PCM-10", "PCM-11"],  # Próximos de M9
    "M1M7": ["PCM-01", "PCM-07"],  # M1/M7 próximos
    "PDO": ["PDO-01", "PDO-02", "PDO-03"],  # PDO cluster
    "PGA": ["PGA-01", "PGA-02", "PGA-03", "PGA-04", "PGA-05", "PGA-07", "PGA-08"],
    "PRB": ["PRB-01"],  # Isolado
}

# Pares obrigatórios: quando ambos têm demanda, devem ir no mesmo barco
MANDATORY_PAIRS = [
    ("PCM-02", "PCM-03"),  # M2+M3: 1.04 NM apart
    ("PCM-06", "PCB-01"),  # M6+B1: 1.48 NM apart
]

# Plataformas que combinam bem em rotas diretas de TMIB
DIRECT_COMPATIBLE = {
    "PCM-06": ["PCB-01", "PCB-02", "PCB-03", "PCB-04", "PCM-08"],  # M6 com B cluster
    "PCB-01": ["PCM-06", "PCB-02", "PCB-03", "PCB-04", "PCM-08"],
    "PCB-02": ["PCM-06", "PCB-01", "PCB-03", "PCB-04", "PCM-08"],
    "PCB-03": ["PCM-06", "PCB-01", "PCB-02", "PCB-04", "PCM-08"],
    "PCB-04": ["PCM-06", "PCB-01", "PCB-02", "PCB-03", "PCM-08"],
    "PCM-02": ["PCM-03", "PCM-10"],  # M2 com M3
    "PCM-03": ["PCM-02", "PCM-10"],
    "PDO-01": ["PDO-02", "PDO-03", "PGA-03", "PGA-04"],  # PDO com PGA próximos
    "PDO-02": ["PDO-01", "PDO-03", "PGA-03", "PGA-08"],
    "PDO-03": ["PDO-01", "PDO-02", "PGA-03", "PGA-08"],
}


# ======== Platform normalization ========

def norm_plat(code: str) -> str:
    c = code.strip().upper()
    if c in ("TMIB", "NORWIND GALE"):
        return c
    if re.match(r"^(PCM|PCB|PGA|PRB|PDO)-\d{2}$", c):
        return c
    m = re.match(r"^M(\d+)$", c)
    if m:
        return f"PCM-{int(m.group(1)):02d}"
    b = re.match(r"^B(\d+)$", c)
    if b:
        return f"PCB-{int(b.group(1)):02d}"
    pg = re.match(r"^PGA(\d+)$", c)
    if pg:
        return f"PGA-{int(pg.group(1)):02d}"
    pdo = re.match(r"^PDO(\d+)$", c)
    if pdo:
        return f"PDO-{int(pdo.group(1)):02d}"
    prb = re.match(r"^PRB(\d+)$", c)
    if prb:
        return f"PRB-{int(prb.group(1)):02d}"
    return c


def short_plat(norm: str) -> str:
    n = norm.upper().strip()
    if n == "TMIB":
        return "TMIB"
    if re.match(r"^PCM-\d{2}$", n):
        return f"M{int(n.split('-')[1])}"
    if re.match(r"^PCB-\d{2}$", n):
        return f"B{int(n.split('-')[1])}"
    if re.match(r"^PGA-\d{2}$", n):
        return f"PGA{int(n.split('-')[1])}"
    if re.match(r"^PDO-\d{2}$", n):
        return f"PDO{int(n.split('-')[1])}"
    if re.match(r"^PRB-\d{2}$", n):
        return f"PRB{int(n.split('-')[1])}"
    return n


def get_geo_cluster(platform_norm: str) -> str:
    for cluster_name, platforms in GEO_CLUSTERS.items():
        if platform_norm in platforms:
            return cluster_name
    return "OTHER"


# ======== Data classes ========

@dataclass
class Demand:
    platform: str
    platform_norm: str
    tmib: int = 0
    m9: int = 0
    priority: int = 99

    def total(self) -> int:
        return self.tmib + self.m9

    def has_m9_demand(self) -> bool:
        return self.m9 > 0

    def copy(self):
        return Demand(self.platform, self.platform_norm, self.tmib, self.m9, self.priority)


@dataclass
class Boat:
    name: str
    available: bool = False
    departure: str = ""
    fixed_route: str = ""
    speed: float = DEFAULT_SPEED_KN
    max_capacity: int = 24

    def departure_minutes(self) -> int:
        if not self.departure or ":" not in self.departure:
            return 999 * 60
        parts = self.departure.split(":")
        return int(parts[0]) * 60 + int(parts[1])


@dataclass
class Config:
    troca_turma: bool = False
    rendidos_m9: int = 0


@dataclass
class Route:
    """Representa uma rota completa de um barco."""
    boat: Boat
    stops: List[Tuple[str, int, int]]  # Paradas pós-M9: [(platform_norm, tmib_drop, m9_drop), ...]
    pre_m9_stops: List[Tuple[str, int, int]] = field(default_factory=list)  # Paradas pré-M9 (TMIB-only)
    m9_pickup: int = 0  # pax M9 embarcados em M9
    tmib_to_m9: int = 0  # pax TMIB desembarcados em M9
    uses_m9_hub: bool = False
    total_distance: float = 0.0
    priority_map: Dict[str, int] = field(default_factory=dict)
    m9_priority: int = 99

    def total_tmib(self) -> int:
        return self.tmib_to_m9 + sum(s[1] for s in self.pre_m9_stops) + sum(s[1] for s in self.stops)

    def total_m9(self) -> int:
        return sum(s[2] for s in self.stops)

    def total_pax(self) -> int:
        return self.max_load()

    def max_load(self) -> int:
        if not self.uses_m9_hub:
            return self.total_tmib()
        pre_load = self.total_tmib()
        post_load = (self.total_tmib() - self.tmib_to_m9) + self.m9_pickup
        return max(pre_load, post_load)


# ======== Load functions ========

def load_distances(path: str) -> Dict[str, Dict[str, float]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    dist = {}
    for a, m in data.items():
        aN = norm_plat(str(a))
        dist.setdefault(aN, {})
        for b, v in m.items():
            dist[aN][norm_plat(str(b))] = float(v)
    return dist


def get_dist(distances, a_norm, b_norm) -> float:
    if a_norm == b_norm:
        return 0.0
    if a_norm in distances and b_norm in distances[a_norm]:
        return distances[a_norm][b_norm]
    if b_norm in distances and a_norm in distances[b_norm]:
        return distances[b_norm][a_norm]
    return 999.0


def load_speeds(path: str) -> Dict[str, float]:
    speeds = {}
    if not os.path.exists(path):
        return speeds
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    section = None
    for line in content.split("\n"):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        sm = re.match(r"\[([A-Z_]+)\]", line)
        if sm:
            section = sm.group(1).replace("_", " ")
            continue
        if "=" in line:
            parts = line.split("=")
            if len(parts) == 2:
                left = parts[0].strip()
                try:
                    spd = float(parts[1].strip())
                except ValueError:
                    continue
                if section and left.isdigit():
                    speeds[f"{section} {left}".upper()] = spd
                else:
                    name = left.replace("_", " ").upper()
                    speeds[name] = spd
                    speeds[left.upper()] = spd
    return speeds


def get_speed(speeds, name) -> float:
    up = name.upper()
    for v in [up, up.replace("_", " "), up.replace(" ", "_")]:
        if v in speeds:
            return speeds[v]
    return DEFAULT_SPEED_KN


def get_max_capacity(name) -> int:
    up = name.upper()
    if "AQUA" in up and "HELIX" in up:
        return 100
    return 24


def is_aqua_helix(name) -> bool:
    up = name.upper()
    return "AQUA" in up and "HELIX" in up


def load_gangway(path: str) -> Set[str]:
    if not os.path.exists(path):
        return set()
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    platforms = data.get("plataformas_gangway", [])
    return {norm_plat(p) for p in platforms}


# ======== Time and distance calculations ========

def travel_time_minutes(distance_nm: float, speed_kn: float) -> int:
    if speed_kn <= 0:
        return 999
    return math.ceil(distance_nm / speed_kn * 60)


def calc_route_distance(route: Route, distances: Dict) -> float:
    """Calcula distância total da rota."""
    total = 0.0
    current = "TMIB"

    if route.uses_m9_hub:
        # Paradas pré-M9
        for stop in route.pre_m9_stops:
            total += get_dist(distances, current, stop[0])
            current = stop[0]

        m9 = norm_plat("M9")
        total += get_dist(distances, current, m9)
        current = m9

    for stop in route.stops:
        total += get_dist(distances, current, stop[0])
        current = stop[0]

    return total


def calc_arrival_times(route: Route, distances: Dict) -> List[Tuple[str, int]]:
    """Calcula horário de chegada em cada plataforma."""
    arrivals = []
    current_time = route.boat.departure_minutes()
    current_pos = "TMIB"
    is_aqua = is_aqua_helix(route.boat.name)

    if route.uses_m9_hub:
        for stop in route.pre_m9_stops:
            dist = get_dist(distances, current_pos, stop[0])
            current_time += travel_time_minutes(dist, route.boat.speed)
            if is_aqua:
                current_time += AQUA_APPROACH_TIME
            arrivals.append((stop[0], current_time))
            current_time += (stop[1] + stop[2]) * MINUTES_PER_PAX
            current_pos = stop[0]

        m9 = norm_plat("M9")
        dist = get_dist(distances, current_pos, m9)
        current_time += travel_time_minutes(dist, route.boat.speed)
        if is_aqua:
            current_time += AQUA_APPROACH_TIME
        # Operação em M9
        current_time += (route.tmib_to_m9 + route.m9_pickup) * MINUTES_PER_PAX
        arrivals.append((m9, current_time))
        current_pos = m9

    for stop in route.stops:
        dist = get_dist(distances, current_pos, stop[0])
        current_time += travel_time_minutes(dist, route.boat.speed)
        if is_aqua:
            current_time += AQUA_APPROACH_TIME
        arrivals.append((stop[0], current_time))
        # Operação
        current_time += (stop[1] + stop[2]) * MINUTES_PER_PAX
        current_pos = stop[0]

    return arrivals


def calc_weighted_arrival_score(route: Route, distances: Dict) -> float:
    """
    Calcula score ponderado de chegada (menor = melhor).
    Prioriza chegada cedo com mais passageiros.
    """
    arrivals = calc_arrival_times(route, distances)
    score = 0.0

    for i, (plat, arrival_min) in enumerate(arrivals):
        if plat == norm_plat("M9"):
            continue  # M9 é hub, não destino final

        # Encontrar quantidade de pax nesta parada
        all_stops = route.pre_m9_stops + route.stops
        for stop in all_stops:
            if stop[0] == plat:
                pax = stop[1] + stop[2]
                # Score: tempo * pax (entregar mais pax cedo é melhor)
                score += arrival_min * pax
                break

    return score


def calc_priority_time_penalty(route: Route, distances: Dict,
                               priority_map: Dict[str, int],
                               m9_priority: int = 99) -> float:
    """
    Penaliza chegada tardia em plataformas prioritárias.
    Usa o menor horário de chegada por plataforma.
    """
    arrivals = calc_arrival_times(route, distances)
    min_arrival: Dict[str, int] = {}
    for plat, arrival_min in arrivals:
        if plat not in min_arrival or arrival_min < min_arrival[plat]:
            min_arrival[plat] = arrival_min

    def weight(priority: int) -> int:
        if priority == 1:
            return 15
        if priority == 2:
            return 3
        if priority == 3:
            return 1
        return 0

    penalty = 0.0
    for plat_norm, arrival_min in min_arrival.items():
        if plat_norm == norm_plat("M9"):
            # M9 priority is driven by TMIB->M9 delivery demand.
            # If there is no TMIB->M9 drop in this route, ignore M9 priority here.
            if route.tmib_to_m9 <= 0:
                continue
            p = m9_priority
        else:
            p = priority_map.get(plat_norm, 99)
        w = weight(p)
        if w > 0:
            penalty += arrival_min * w

    return penalty


def calc_comfort_pax_minutes(route: Route, distances: Dict) -> float:
    """
    Calcula pax-minuto a bordo (menor = melhor).
    Considera tempos de deslocamento, aproximação (Aqua) e operação.
    """
    tmib_onboard = route.total_tmib()
    m9_onboard = 0
    total = 0.0
    current = "TMIB"
    is_aqua = is_aqua_helix(route.boat.name)

    def add_segment_time(minutes: int):
        nonlocal total
        total += (tmib_onboard + m9_onboard) * minutes

    def travel_to(dest: str):
        dist = get_dist(distances, current, dest)
        travel = travel_time_minutes(dist, route.boat.speed)
        add_segment_time(travel)
        if is_aqua:
            add_segment_time(AQUA_APPROACH_TIME)
        return travel

    def operate(tmib_drop: int, m9_drop: int, m9_pick: int):
        nonlocal tmib_onboard, m9_onboard
        ops = (tmib_drop + m9_drop + m9_pick) * MINUTES_PER_PAX
        add_segment_time(ops)
        tmib_onboard -= tmib_drop
        m9_onboard -= m9_drop
        m9_onboard += m9_pick

    if route.uses_m9_hub:
        for stop in route.pre_m9_stops:
            dest = stop[0]
            travel_to(dest)
            operate(stop[1], stop[2], 0)
            current = dest

        m9 = norm_plat("M9")
        travel_to(m9)
        operate(route.tmib_to_m9, 0, route.m9_pickup)
        current = m9

        for stop in route.stops:
            dest = stop[0]
            travel_to(dest)
            operate(stop[1], stop[2], 0)
            current = dest
    else:
        for stop in route.stops:
            dest = stop[0]
            travel_to(dest)
            operate(stop[1], stop[2], 0)
            current = dest

    return total


def calc_cluster_cohesion_penalty(route: Route, distances: Dict) -> float:
    """
    Penaliza rotas com baixa coesao geografica.
    Objetivo: privilegiar blocos de plataformas proximas no mesmo barco.
    """
    def segment_penalty(stops: List[Tuple[str, int, int]]) -> float:
        penalty = 0.0
        prev_cluster = None
        prev_plat = None

        for plat_norm, _, _ in stops:
            cluster = get_geo_cluster(plat_norm)
            if prev_cluster is not None and cluster != prev_cluster:
                if are_clusters_compatible(prev_cluster, cluster):
                    penalty += CLUSTER_SWITCH_PENALTY_NM
                else:
                    penalty += INCOMPATIBLE_CLUSTER_SWITCH_PENALTY_NM

                jump_nm = get_dist(distances, prev_plat, plat_norm)
                excess_nm = max(0.0, jump_nm - CROSS_CLUSTER_JUMP_FREE_NM)
                penalty += excess_nm * CROSS_CLUSTER_JUMP_PENALTY_PER_NM

            prev_cluster = cluster
            prev_plat = plat_norm

        return penalty

    return segment_penalty(route.pre_m9_stops) + segment_penalty(route.stops)


def order_stops_with_priority(stops: List[Tuple[str, int, int]],
                              distances: Dict,
                              start: str,
                              boat: Boat,
                              priority_map: Dict[str, int]) -> List[Tuple[str, int, int]]:
    """
    Ordena paradas considerando distância, prioridades e pax entregues cedo.
    Para conjuntos pequenos, testa todas as permutações; caso contrário, usa heurística gulosa.
    """
    if len(stops) <= 1:
        return list(stops)

    has_priority = any(priority_map.get(s[0], 99) <= 3 for s in stops)
    if not has_priority:
        # fallback para distância pura
        stop_demands = [Demand(short_plat(s[0]), s[0], s[1], s[2]) for s in stops]
        ordered = optimal_order_from(stop_demands, distances, start)
        return [(d.platform_norm, d.tmib, d.m9) for d in ordered]

    is_aqua = is_aqua_helix(boat.name)

    def weight(priority: int) -> int:
        if priority == 1:
            return 15
        if priority == 2:
            return 3
        if priority == 3:
            return 1
        return 0

    def score(order: Tuple[Tuple[str, int, int], ...]) -> float:
        current = start
        dist_total = 0.0
        time = 0
        score_priority = 0.0
        score_pax = 0.0
        comfort = 0.0
        backtrack = 0.0
        p1_precedence_penalty = 0.0
        onboard = sum(s[1] + s[2] for s in order)
        prev_radial = None
        remaining_priority1 = sum(1 for s in order if priority_map.get(s[0], 99) == 1)

        for stop in order:
            p = priority_map.get(stop[0], 99)
            if p != 1 and remaining_priority1 > 0:
                # Quase-hard: evite colocar prioridade 1 para depois.
                p1_precedence_penalty += PRIORITY1_PRECEDENCE_PENALTY_NM
            if p == 1:
                remaining_priority1 -= 1

            dist = get_dist(distances, current, stop[0])
            dist_total += dist

            travel = travel_time_minutes(dist, boat.speed)
            segment = travel + (AQUA_APPROACH_TIME if is_aqua else 0)
            comfort += onboard * segment
            time += segment

            pax = stop[1] + stop[2]
            score_pax += time * pax
            score_priority += time * weight(p)

            ops = pax * MINUTES_PER_PAX
            comfort += onboard * ops
            time += ops
            onboard -= pax
            current = stop[0]

            radial = get_dist(distances, start, stop[0])
            if prev_radial is not None and radial < prev_radial:
                backtrack += (prev_radial - radial)
            prev_radial = radial

        return (
            dist_total
            + (score_priority * PRIORITY_TIME_WEIGHT)
            + (score_pax * PAX_ARRIVAL_WEIGHT)
            + (comfort * COMFORT_PAX_MIN_WEIGHT)
            + (backtrack * BACKTRACK_PENALTY_NM)
            + p1_precedence_penalty
        )

    if len(stops) <= 7:
        from itertools import permutations
        best = min(permutations(stops), key=score)
        return list(best)

    # Heurística gulosa para conjuntos grandes
    remaining = list(stops)
    ordered = []
    current = start
    while remaining:
        best = min(
            remaining,
            key=lambda s: score(tuple([s]))
        )
        ordered.append(best)
        current = best[0]
        remaining.remove(best)
    return ordered


# ======== Fixed route parser ========

def parse_fixed_route(route_str: str) -> Dict[str, Dict[str, int]]:
    """Parseia uma rota fixa e retorna as entregas por plataforma."""
    deliveries = {}
    parts = route_str.split('/')

    for part in parts:
        part = part.strip()
        if not part:
            continue

        tokens = part.split()
        platform = tokens[0]
        platform_norm = norm_plat(platform)

        if platform_norm == 'TMIB':
            continue

        tmib_drop = 0
        m9_drop = 0

        for token in tokens[1:]:
            token = token.strip()
            if re.match(r'^\(-\d+\)$', token):
                m9_drop += int(token[2:-1])
            elif re.match(r'^-\d+$', token):
                tmib_drop += int(token[1:])

        if tmib_drop > 0 or m9_drop > 0:
            if platform_norm not in deliveries:
                deliveries[platform_norm] = {'tmib': 0, 'm9': 0}
            deliveries[platform_norm]['tmib'] += tmib_drop
            deliveries[platform_norm]['m9'] += m9_drop

    return deliveries


# ======== Read input ========

def read_solver_input(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    config = Config()
    troca = ws.cell(row=4, column=3).value
    config.troca_turma = str(troca).strip().upper() == "SIM" if troca else False
    rend = ws.cell(row=5, column=3).value
    config.rendidos_m9 = int(rend) if rend else 0

    boats = []
    r = 9
    while True:
        name = ws.cell(row=r, column=2).value
        if not name or str(name).strip() == "":
            break
        disp = ws.cell(row=r, column=3).value
        hora = ws.cell(row=r, column=4).value
        rota = ws.cell(row=r, column=5).value

        if hora and hasattr(hora, 'strftime'):
            hora = hora.strftime("%H:%M")
        elif hora and ":" not in str(hora):
            try:
                h_val = float(hora) * 24
                hora = f"{int(h_val):02d}:{int((h_val % 1) * 60):02d}"
            except:
                hora = ""
        else:
            hora = str(hora).strip() if hora else ""

        rota_str = str(rota).strip() if rota and str(rota).strip().upper() != "NONE" else ""

        boats.append(Boat(
            name=str(name).strip(),
            available=str(disp).strip().upper() == "SIM" if disp else False,
            departure=hora,
            fixed_route=rota_str,
            max_capacity=get_max_capacity(str(name).strip()),
        ))
        r += 1

    # Demand section
    demand_header_row = r + 1
    demand_col_row = demand_header_row + 1
    demand_start = demand_col_row + 1

    demands = []
    r = demand_start
    while True:
        plat = ws.cell(row=r, column=2).value
        if not plat or str(plat).strip() == "":
            break
        plat_str = str(plat).strip()
        m9_val = ws.cell(row=r, column=3).value
        tmib_val = ws.cell(row=r, column=4).value
        prio_val = ws.cell(row=r, column=5).value

        demands.append(Demand(
            platform=plat_str,
            platform_norm=norm_plat(plat_str),
            tmib=int(tmib_val) if tmib_val else 0,
            m9=int(m9_val) if m9_val else 0,
            priority=int(prio_val) if prio_val else 99,
        ))
        r += 1

    return config, boats, demands


# ======== Routing algorithms ========

def nn_order_from(platforms: List[Demand], distances: Dict, start: str) -> List[Demand]:
    """Ordena plataformas por vizinho mais próximo a partir de start."""
    if len(platforms) <= 1:
        return list(platforms)

    remaining = list(platforms)
    ordered = []
    current = start

    while remaining:
        best = min(remaining, key=lambda p: get_dist(distances, current, p.platform_norm))
        ordered.append(best)
        current = best.platform_norm
        remaining.remove(best)

    return ordered


def optimal_order_from(platforms: List[Demand], distances: Dict, start: str) -> List[Demand]:
    """
    Para conjuntos pequenos (≤6 paradas), testa TODAS as permutações e
    retorna a de menor distância total. Para conjuntos maiores, usa NN.
    """
    from itertools import permutations

    if len(platforms) <= 1:
        return list(platforms)

    if len(platforms) > 6:
        return nn_order_from(platforms, distances, start)

    best_order = None
    best_dist = float('inf')

    for perm in permutations(platforms):
        total = 0.0
        current = start
        for p in perm:
            total += get_dist(distances, current, p.platform_norm)
            current = p.platform_norm
        if total < best_dist:
            best_dist = total
            best_order = list(perm)

    return best_order


def split_pre_m9_stops(stops: List[Tuple[str, int, int]],
                       m9_pickup: int,
                       distances: Dict,
                       cap: int) -> Optional[Tuple[List[Tuple[str, int, int]], List[Tuple[str, int, int]]]]:
    """
    Decide quais paradas TMIB-only devem ocorrer ANTES de M9 para respeitar capacidade.
    Retorna (pre_m9_stops, post_m9_stops). Se inviável, retorna None.
    """
    total_tmib = sum(s[1] for s in stops)
    post_load = total_tmib + m9_pickup
    if post_load <= cap:
        return [], list(stops)

    needed = post_load - cap
    candidates = [s for s in stops if s[1] > 0]
    if sum(s[1] for s in candidates) < needed:
        return None

    m9 = norm_plat("M9")
    def detour_cost(stop):
        plat = stop[0]
        return get_dist(distances, "TMIB", plat) + get_dist(distances, plat, m9) - get_dist(distances, "TMIB", m9)

    # Escolher paradas com menor desvio para ante-M9
    candidates_sorted = sorted(
        candidates,
        key=lambda s: (1 if s[2] > 0 else 0, detour_cost(s), -s[1])
    )
    moved = 0
    pre_m9 = []

    post_m9 = [list(s) for s in stops]
    for s in candidates_sorted:
        if moved >= needed:
            break
        # mover toda a parte TMIB desta parada para antes de M9
        pre_m9.append((s[0], s[1], 0))
        moved += s[1]
        # zerar TMIB no pós-M9
        for p in post_m9:
            if p[0] == s[0]:
                p[1] = 0
                break

    post_m9 = [(p[0], p[1], p[2]) for p in post_m9 if p[1] > 0 or p[2] > 0]
    return pre_m9, post_m9


def promote_priority1_pre_m9(pre_m9_stops: List[Tuple[str, int, int]],
                             post_m9_stops: List[Tuple[str, int, int]],
                             distances: Dict,
                             priority_map: Dict[str, int]) -> Tuple[List[Tuple[str, int, int]], List[Tuple[str, int, int]]]:
    """
    Promove paradas prioridade 1 (TMIB-only) para antes de M9 quando o desvio for pequeno.
    Isso aproxima um comportamento quase-hard sem forcar saltos longos pre-M9.
    """
    if not post_m9_stops:
        return pre_m9_stops, post_m9_stops

    m9 = norm_plat("M9")
    moved = []
    kept = []

    for stop in post_m9_stops:
        plat, tmib_drop, m9_drop = stop
        if tmib_drop > 0 and m9_drop == 0 and priority_map.get(plat, 99) == 1:
            detour = (
                get_dist(distances, "TMIB", plat)
                + get_dist(distances, plat, m9)
                - get_dist(distances, "TMIB", m9)
            )
            if detour <= PRIORITY1_PRE_M9_MAX_DETOUR_NM:
                moved.append(stop)
                continue
        kept.append(stop)

    if not moved:
        return pre_m9_stops, post_m9_stops

    return pre_m9_stops + moved, kept


def rebuild_pre_m9(route: Route, distances: Dict) -> bool:
    """
    Recalcula a divisão pré/pós-M9 para uma rota existente.
    Retorna True se a rota permanece viável.
    """
    if not route.uses_m9_hub:
        route.pre_m9_stops = []
        return True

    all_stops = list(route.pre_m9_stops) + list(route.stops)
    total_tmib = sum(s[1] for s in all_stops)
    total_m9 = sum(s[2] for s in all_stops)

    pre_load = total_tmib + route.tmib_to_m9
    if pre_load > route.boat.max_capacity:
        return False

    split = split_pre_m9_stops(all_stops, total_m9, distances, route.boat.max_capacity)
    if split is None:
        return False

    pre_m9_stops, post_m9_stops = split
    route.pre_m9_stops = pre_m9_stops
    route.stops = post_m9_stops
    route.m9_pickup = total_m9
    route.uses_m9_hub = route.m9_pickup > 0 or route.tmib_to_m9 > 0
    return True


def find_compatible_platforms(platform_norm: str, demands: List[Demand]) -> List[Demand]:
    """Encontra plataformas compatíveis para rota direta."""
    compatible = DIRECT_COMPATIBLE.get(platform_norm, [])
    return [d for d in demands if d.platform_norm in compatible]


def build_direct_route(boat: Boat, demands: List[Demand], distances: Dict) -> Optional[Route]:
    """
    Constrói rota direta TMIB→plataformas (sem parar em M9).
    Só funciona para plataformas sem demanda M9.
    """
    # Filtrar apenas demandas TMIB-only
    tmib_only = [d for d in demands if d.tmib > 0 and d.m9 == 0]
    if not tmib_only:
        return None

    cap = boat.max_capacity

    # Tentar construir rota começando pela plataforma mais próxima de TMIB
    tmib_only.sort(key=lambda d: get_dist(distances, "TMIB", d.platform_norm))

    best_route = None
    best_score = float('inf')
    priority_map = {d.platform_norm: d.priority for d in tmib_only}

    # Tentar cada plataforma como ponto inicial
    for start_demand in tmib_only:
        stops = []
        total_pax = 0
        used = set()

        # Adicionar plataforma inicial
        if start_demand.tmib <= cap:
            stops.append((start_demand.platform_norm, start_demand.tmib, 0))
            total_pax = start_demand.tmib
            used.add(start_demand.platform_norm)

            # Encontrar plataformas compatíveis
            current = start_demand.platform_norm
            compatible = find_compatible_platforms(current, tmib_only)

            # Ordenar por distância
            compatible = [d for d in compatible if d.platform_norm not in used]
            compatible.sort(key=lambda d: get_dist(distances, current, d.platform_norm))

            for d in compatible:
                if total_pax + d.tmib <= cap:
                    stops.append((d.platform_norm, d.tmib, 0))
                    total_pax += d.tmib
                    used.add(d.platform_norm)
                    current = d.platform_norm

        if stops:
            stops = order_stops_with_priority(stops, distances, "TMIB", boat, priority_map)
            route = Route(
                boat=boat,
                stops=stops,
                m9_pickup=0,
                tmib_to_m9=0,
                uses_m9_hub=False,
                priority_map=priority_map,
            )
            route.total_distance = calc_route_distance(route, distances)
            score = calc_weighted_arrival_score(route, distances)

            if score < best_score:
                best_score = score
                best_route = route

    return best_route


def build_m9_hub_route(boat: Boat, demands: List[Demand], m9_tmib_demand: int,
                       distances: Dict, gangway_platforms: Set[str],
                       target_cluster: str = None) -> Optional[Route]:
    """
    Constrói rota via M9 hub.
    Se target_cluster especificado, prioriza plataformas desse cluster.
    """
    cap = boat.max_capacity
    is_aqua = is_aqua_helix(boat.name)
    m9_norm = norm_plat("M9")

    # Filtrar plataformas que este barco pode servir
    if is_aqua:
        servable = [d for d in demands if d.platform_norm in gangway_platforms]
    else:
        servable = list(demands)

    if not servable:
        return None

    # Se target_cluster especificado, filtrar para esse cluster
    if target_cluster:
        cluster_demands = [d for d in servable if get_geo_cluster(d.platform_norm) == target_cluster]
        if cluster_demands:
            servable = cluster_demands

    # Separar por tipo de demanda
    m9_demand_platforms = [d for d in servable if d.m9 > 0]
    tmib_only_platforms = [d for d in servable if d.tmib > 0 and d.m9 == 0]
    priority_map = {d.platform_norm: d.priority for d in servable}

    # Ordenar por quantidade de M9 (maior primeiro), depois por distância
    m9_demand_platforms.sort(key=lambda d: (-d.m9, get_dist(distances, m9_norm, d.platform_norm)))

    stops = []
    total_tmib = 0
    total_m9_pickup = 0
    tmib_to_m9 = 0
    current_cluster = None

    # Primeiro, alocar plataformas com demanda M9 (respeitando clusters)
    for d in m9_demand_platforms:
        d_cluster = get_geo_cluster(d.platform_norm)

        # Se já temos stops, verificar compatibilidade de cluster
        if stops and current_cluster:
            # Não misturar clusters distantes
            if d_cluster != current_cluster:
                # Exceção: clusters próximos podem ser misturados
                if not are_clusters_compatible(current_cluster, d_cluster):
                    continue

        new_total_tmib = total_tmib + d.tmib
        new_total_m9 = total_m9_pickup + d.m9
        if new_total_tmib <= cap and new_total_m9 <= cap:
            stops.append((d.platform_norm, d.tmib, d.m9))
            total_tmib += d.tmib
            total_m9_pickup += d.m9
            current_cluster = d_cluster

    # Preencher com TMIB para M9 se houver espaço (não afeta carga pós-M9)
    space_for_m9 = cap - total_tmib
    if space_for_m9 > 0 and m9_tmib_demand > 0:
        tmib_to_m9 = min(space_for_m9, m9_tmib_demand)

    # Adicionar plataformas TMIB-only do MESMO cluster se couber
    for d in tmib_only_platforms:
        d_cluster = get_geo_cluster(d.platform_norm)

        # Só adicionar se for do mesmo cluster ou cluster compatível
        if current_cluster and d_cluster != current_cluster:
            if not are_clusters_compatible(current_cluster, d_cluster):
                continue

        if total_tmib + d.tmib <= cap:
            stops.append((d.platform_norm, d.tmib, 0))
            total_tmib += d.tmib

    if not stops and tmib_to_m9 == 0:
        return None

    # Dividir paradas pré/post-M9 para respeitar capacidade pós-M9
    split = split_pre_m9_stops(stops, total_m9_pickup, distances, cap)
    if split is None:
        return None
    pre_m9_stops, post_m9_stops = split
    pre_m9_stops, post_m9_stops = promote_priority1_pre_m9(
        pre_m9_stops, post_m9_stops, distances, priority_map
    )

    # Reordenar pre-M9 a partir de TMIB e post-M9 a partir de M9
    if len(pre_m9_stops) > 1:
        pre_m9_stops = order_stops_with_priority(pre_m9_stops, distances, "TMIB", boat, priority_map)

    if len(post_m9_stops) > 1:
        post_m9_stops = order_stops_with_priority(post_m9_stops, distances, m9_norm, boat, priority_map)

    route = Route(
        boat=boat,
        stops=post_m9_stops,
        pre_m9_stops=pre_m9_stops,
        m9_pickup=total_m9_pickup,
        tmib_to_m9=tmib_to_m9,
        uses_m9_hub=(total_m9_pickup > 0 or tmib_to_m9 > 0),
        priority_map=priority_map,
    )
    route.total_distance = calc_route_distance(route, distances)

    return route


def are_clusters_compatible(cluster1: str, cluster2: str) -> bool:
    """Verifica se dois clusters podem ser combinados em uma rota."""
    if cluster1 == cluster2:
        return True

    # Clusters que podem ser combinados (geograficamente próximos)
    compatible_pairs = [
        ("M6_AREA", "B_CLUSTER"),  # M6/M8 com B1-B4 (muito próximos)
        ("M6_AREA", "M1M7"),  # M6 com M7
        ("M9_NEAR", "M2M3"),  # M2/M3 perto de M9
        ("M2M3", "M1M7"),  # M2/M3 com M1/M7 (razoavelmente próximos)
        ("M2M3", "M6_AREA"),  # M2/M3 com M6 (via M9)
        ("M2M3", "B_CLUSTER"),  # M2/M3 com B cluster (todos perto de M9)
        ("B_CLUSTER", "M1M7"),  # B cluster com M1/M7
        ("PDO", "PGA"),  # Clusters distantes juntos
    ]

    for c1, c2 in compatible_pairs:
        if (cluster1 == c1 and cluster2 == c2) or (cluster1 == c2 and cluster2 == c1):
            return True
    return False


def is_distant_cluster(cluster: str) -> bool:
    """Verifica se é um cluster distante (>10 NM de TMIB em média)."""
    return cluster in ["PDO", "PGA", "PRB"]


def build_cluster_route(boat: Boat, cluster_demands: List[Demand],
                        m9_tmib_demand: int, distances: Dict,
                        needs_m9_stop: bool) -> Optional[Route]:
    """
    Constrói rota otimizada para um cluster específico.
    """
    cap = boat.max_capacity

    priority_map = {d.platform_norm: d.priority for d in cluster_demands}
    if needs_m9_stop:
        start = norm_plat("M9")
    else:
        start = "TMIB"

    ordered_tuples = order_stops_with_priority(
        [(d.platform_norm, d.tmib, d.m9) for d in cluster_demands],
        distances,
        start,
        boat,
        priority_map,
    )
    ordered = [Demand(short_plat(t[0]), t[0], t[1], t[2]) for t in ordered_tuples]

    stops = []
    total_tmib = 0
    total_m9 = 0
    tmib_to_m9 = 0

    for d in ordered:
        if total_tmib + d.tmib <= cap and total_m9 + d.m9 <= cap:
            stops.append((d.platform_norm, d.tmib, d.m9))
            total_tmib += d.tmib
            total_m9 += d.m9

    # Se usa M9 hub, calcular TMIB para M9
    if needs_m9_stop:
        space = cap - total_tmib
        if space > 0 and m9_tmib_demand > 0:
            tmib_to_m9 = min(space, m9_tmib_demand)

    if not stops:
        return None

    pre_m9_stops = []
    post_m9_stops = stops
    if needs_m9_stop:
        split = split_pre_m9_stops(stops, total_m9, distances, cap)
        if split is None:
            return None
        pre_m9_stops, post_m9_stops = split
        pre_m9_stops, post_m9_stops = promote_priority1_pre_m9(
            pre_m9_stops, post_m9_stops, distances, priority_map
        )

    route = Route(
        boat=boat,
        stops=post_m9_stops,
        pre_m9_stops=pre_m9_stops,
        m9_pickup=total_m9,
        tmib_to_m9=tmib_to_m9,
        uses_m9_hub=needs_m9_stop,
        priority_map=priority_map,
    )
    route.total_distance = calc_route_distance(route, distances)

    return route


# ======== Combinatorial optimizer ========

def form_demand_packages(demands: List[Demand], boats: List[Boat]) -> List[List[Demand]]:
    """
    Agrupa demandas em pacotes baseados em pares obrigatórios.
    Quando ambas as plataformas de um par têm demanda, são agrupadas como unidade atômica.
    """
    packages = []
    used = set()
    max_capacity = max((b.max_capacity for b in boats), default=0)
    n_boats = len(boats)

    for p1, p2 in MANDATORY_PAIRS:
        d1 = next((d for d in demands if d.platform_norm == p1 and d.total() > 0), None)
        d2 = next((d for d in demands if d.platform_norm == p2 and d.total() > 0), None)
        if d1 and d2:
            # Só manter o par se couber em algum barco (pre-load)
            if d1.tmib + d2.tmib <= max_capacity:
                packages.append([d1, d2])
                used.add(p1)
                used.add(p2)

    # Opcional: em cenarios apertados (<=2 barcos), permitir split de UMA
    # demanda TMIB-only grande para destravar melhor agrupamento geografico.
    split_candidate = None
    if n_boats <= 2:
        unsplitted = [
            d for d in demands
            if d.platform_norm not in used and d.m9 == 0 and d.tmib >= 12
        ]
        if unsplitted:
            # Preferir plataformas do entorno de M9/M2-M3 para melhorar
            # combinacoes de bloco com clusters proximos.
            def split_rank(d: Demand):
                c = get_geo_cluster(d.platform_norm)
                pref = 0 if c in ("M2M3", "M9_NEAR") else 1
                return (pref, -d.tmib)

            split_candidate = min(unsplitted, key=split_rank)

    for d in demands:
        if d.platform_norm in used or d.total() <= 0:
            continue

        if split_candidate and d.platform_norm == split_candidate.platform_norm:
            first_chunk = 4
            second_chunk = d.tmib - first_chunk
            if second_chunk > 0:
                packages.append([
                    Demand(d.platform, d.platform_norm, tmib=first_chunk, m9=0, priority=d.priority)
                ])
                packages.append([
                    Demand(d.platform, d.platform_norm, tmib=second_chunk, m9=0, priority=d.priority)
                ])
                continue

        packages.append([d])

    return packages


def evaluate_boat_route(boat_demands: List[Demand], boat: Boat,
                        distances: Dict, m9_tmib_avail: int,
                        gangway_platforms: Set[str],
                        m9_priority: int = 99) -> Tuple[Optional[Route], float, int, float, float, float, float]:
    """
    Constrói e avalia rota para um barco com demandas específicas.
    Retorna (route, distance, m9_tmib_used).
    distance = inf indica atribuição inválida.
    """
    if not boat_demands:
        return None, 0.0, 0, 0.0, 0.0, 0.0, 0.0

    # Consolidar eventuais splits da mesma plataforma no mesmo barco.
    merged: Dict[str, Demand] = {}
    for d in boat_demands:
        if d.platform_norm not in merged:
            merged[d.platform_norm] = d.copy()
        else:
            md = merged[d.platform_norm]
            md.tmib += d.tmib
            md.m9 += d.m9
            md.priority = min(md.priority, d.priority)
    boat_demands = list(merged.values())

    is_aqua = is_aqua_helix(boat.name)
    cap = boat.max_capacity
    m9_norm = norm_plat("M9")

    # Verificar restrição de gangway para Aqua
    if is_aqua:
        if any(d.platform_norm not in gangway_platforms for d in boat_demands):
            return None, float('inf'), 0, 0.0, 0.0, 0.0, 0.0

    total_m9_pickup = sum(d.m9 for d in boat_demands)
    total_tmib_deliver = sum(d.tmib for d in boat_demands)
    needs_m9 = total_m9_pickup > 0

    # Calcular TMIB→M9 delivery (preencher espaço livre pré-M9).
    # Regra: permitir parada em M9 mesmo sem embarque M9->plataforma,
    # para não deixar TMIB->M9 pendente quando há capacidade disponível.
    space = cap - total_tmib_deliver
    tmib_to_m9 = 0
    if space > 0 and m9_tmib_avail > 0:
        tmib_to_m9 = min(space, m9_tmib_avail)
        needs_m9 = True

    pre_load = total_tmib_deliver + tmib_to_m9
    # A carga inicial (pre-M9) nunca pode exceder a capacidade.
    # A carga pos-M9 e validada por split_pre_m9_stops, que move
    # desembarques TMIB para antes de M9 quando necessario.
    if pre_load > cap:
        return None, float('inf'), 0, 0.0, 0.0, 0.0, 0.0

    stops = [(d.platform_norm, d.tmib, d.m9) for d in boat_demands]

    priority_map = {d.platform_norm: d.priority for d in boat_demands}
    pre_m9_stops = []
    post_m9_stops = stops
    if needs_m9:
        split = split_pre_m9_stops(stops, total_m9_pickup, distances, cap)
        if split is None:
            return None, float('inf'), 0, 0.0, 0.0, 0.0, 0.0
        pre_m9_stops, post_m9_stops = split
        pre_m9_stops, post_m9_stops = promote_priority1_pre_m9(
            pre_m9_stops, post_m9_stops, distances, priority_map
        )
    start = m9_norm if needs_m9 else "TMIB"
    if needs_m9 and len(pre_m9_stops) > 1:
        pre_m9_stops = order_stops_with_priority(pre_m9_stops, distances, "TMIB", boat, priority_map)

    if len(post_m9_stops) > 1:
        post_m9_stops = order_stops_with_priority(post_m9_stops, distances, start, boat, priority_map)

    route = Route(
        boat=boat,
        stops=post_m9_stops,
        pre_m9_stops=pre_m9_stops,
        m9_pickup=total_m9_pickup,
        tmib_to_m9=tmib_to_m9,
        uses_m9_hub=(total_m9_pickup > 0 or tmib_to_m9 > 0),
        priority_map=priority_map,
        m9_priority=m9_priority,
    )
    route.total_distance = calc_route_distance(route, distances)
    priority_penalty = calc_priority_time_penalty(route, distances, priority_map, m9_priority)
    comfort_cost = calc_comfort_pax_minutes(route, distances)
    pax_arrival_score = calc_weighted_arrival_score(route, distances)
    cluster_penalty = calc_cluster_cohesion_penalty(route, distances)

    return route, route.total_distance, tmib_to_m9, priority_penalty, comfort_cost, pax_arrival_score, cluster_penalty


def optimize_hub_assignments(packages: List[List[Demand]], boats: List[Boat],
                              distances: Dict, m9_tmib_demand: int,
                              gangway_platforms: Set[str],
                              m9_priority: int = 99,
                              distant_boats_already: int = 0,
                              max_distant_boats: int = 1) -> Tuple[List[Route], int]:
    """
    Otimização combinatória: tenta todas as atribuições válidas de pacotes
    a barcos e retorna a de menor distância total.
    Aplica penalidade se os embarques M9 ficarem espalhados em varios barcos.

    Escala: com 5 pacotes e 3 barcos = 243 combinações (muito rápido).
    """
    from itertools import product as iter_product

    n_pkgs = len(packages)
    n_boats = len(boats)

    if n_pkgs == 0 or n_boats == 0:
        return [], m9_tmib_demand

    def route_has_distant(route: Route) -> bool:
        for stop in route.pre_m9_stops + route.stops:
            if is_distant_cluster(get_geo_cluster(stop[0])):
                return True
        return False

    def run_optimizer(enforce_all: bool, enforce_distant: bool, require_zero_m9: bool):
        best_routes = None
        best_score = float('inf')
        best_m9_remaining = m9_tmib_demand
        top_n = 5
        top_scores = []  # list of tuples (remaining_m9, score, dist, penalty, priority, comfort, pax_arrival, cluster, assignment)

        for assignment in iter_product(range(n_boats), repeat=n_pkgs):
            # Agrupar pacotes por barco
            boat_demands_map: Dict[int, List[Demand]] = {i: [] for i in range(n_boats)}
            for pkg_idx, boat_idx in enumerate(assignment):
                boat_demands_map[boat_idx].extend(packages[pkg_idx])

            if enforce_all and any(len(boat_demands_map[i]) == 0 for i in range(n_boats)):
                continue

            # Avaliar esta atribuicao
            routes = []
            route_by_boat_idx: Dict[int, Route] = {}
            total_dist = 0.0
            total_priority_penalty = 0.0
            total_comfort_cost = 0.0
            total_pax_arrival_score = 0.0
            total_cluster_penalty = 0.0
            remaining_m9 = m9_tmib_demand
            valid = True

            for boat_idx in range(n_boats):
                demands_for_boat = boat_demands_map[boat_idx]
                if not demands_for_boat:
                    continue

                route, dist, m9_used, priority_penalty, comfort_cost, pax_arrival_score, cluster_penalty = evaluate_boat_route(
                    demands_for_boat, boats[boat_idx], distances,
                    remaining_m9, gangway_platforms, m9_priority
                )

                if dist == float('inf'):
                    valid = False
                    break

                if route:
                    routes.append(route)
                    route_by_boat_idx[boat_idx] = route
                    total_dist += dist
                    total_priority_penalty += priority_penalty
                    total_comfort_cost += comfort_cost
                    total_pax_arrival_score += pax_arrival_score
                    total_cluster_penalty += cluster_penalty
                    remaining_m9 -= m9_used

            if valid:
                if require_zero_m9 and remaining_m9 > 0:
                    continue

                # Penalizar distribuicao de pax M9 em varios barcos
                m9_routes = sum(1 for r in routes if r.m9_pickup > 0 or r.tmib_to_m9 > 0)
                penalty = max(0, m9_routes - 1) * M9_CONSOLIDATION_PENALTY_NM

                if enforce_distant:
                    distant_now = sum(1 for r in routes if route_has_distant(r))
                    if distant_boats_already + distant_now > max_distant_boats:
                        valid = False
                        scored_dist = float('inf')
                        continue

                # Regra operacional: se existe P1 e tambem P2/P3,
                # evite separar P2/P3 quando ele caberia em um barco que ja leva P1.
                has_p1 = any(d.priority == 1 for demands_for_boat in boat_demands_map.values() for d in demands_for_boat)
                has_p23 = any(d.priority in (2, 3) for demands_for_boat in boat_demands_map.values() for d in demands_for_boat)
                priority_mix_penalty = 0.0
                if has_p1 and has_p23:
                    p1_boats = {
                        b_idx for b_idx, ds in boat_demands_map.items()
                        if any(d.priority == 1 for d in ds)
                    }
                    p23_items = []
                    for b_idx, ds in boat_demands_map.items():
                        for d in ds:
                            if d.priority in (2, 3):
                                p23_items.append((b_idx, d))

                    for cur_boat_idx, d in p23_items:
                        same_p1_boat = cur_boat_idx in p1_boats
                        if same_p1_boat:
                            continue

                        can_fit_with_some_p1 = False
                        for p1_boat_idx in p1_boats:
                            r = route_by_boat_idx.get(p1_boat_idx)
                            if not r:
                                continue
                            free = r.boat.max_capacity - r.max_load()
                            if free >= d.total():
                                can_fit_with_some_p1 = True
                                break

                        if can_fit_with_some_p1:
                            priority_mix_penalty += PRIORITY_MIX_FIT_PENALTY_NM

                cluster_weight = 1.0 if n_boats <= 2 else 0.0

                scored_dist = (
                    total_dist
                    + penalty
                    + priority_mix_penalty
                    + (total_priority_penalty * PRIORITY_TIME_WEIGHT)
                    + (total_comfort_cost * COMFORT_PAX_MIN_WEIGHT)
                    + (total_pax_arrival_score * PAX_ARRIVAL_WEIGHT)
                    + (total_cluster_penalty * cluster_weight)
                )

                # Guardar top combinacoes
                top_scores.append((
                    remaining_m9,
                    scored_dist,
                    total_dist,
                    penalty,
                    total_priority_penalty,
                    total_comfort_cost,
                    total_pax_arrival_score,
                    total_cluster_penalty,
                    assignment,
                ))
                top_scores.sort(key=lambda x: (x[0], x[1]))
                if len(top_scores) > top_n:
                    top_scores = top_scores[:top_n]
            else:
                scored_dist = float('inf')

            # Objetivo lexicografico:
            # 1) minimizar demanda TMIB->M9 nao atendida
            # 2) entre empates, minimizar score total
            if valid:
                better_on_m9 = remaining_m9 < best_m9_remaining
                tie_on_m9_better_score = remaining_m9 == best_m9_remaining and scored_dist < best_score
                if better_on_m9 or tie_on_m9_better_score:
                    best_score = scored_dist
                    best_routes = routes
                    best_m9_remaining = remaining_m9

        return best_routes, best_m9_remaining, top_scores

    enforce_all = n_pkgs >= n_boats
    enforce_distant = max_distant_boats is not None and max_distant_boats > 0

    def run_with_relaxations(require_zero_m9: bool):
        local_best_routes, local_best_m9_remaining, local_top_scores = run_optimizer(
            enforce_all, enforce_distant, require_zero_m9
        )
        if enforce_all and not local_best_routes:
            if require_zero_m9:
                print("  AVISO: sem solucao com m9_restante=0 usando todos os barcos; relaxando a restricao.")
            else:
                print("  AVISO: nao foi possivel usar todos os barcos; relaxando a restricao.")
            local_best_routes, local_best_m9_remaining, local_top_scores = run_optimizer(
                False, enforce_distant, require_zero_m9
            )

        if enforce_distant and not local_best_routes:
            if require_zero_m9:
                print("  AVISO: sem solucao com m9_restante=0 limitando barcos distantes; relaxando a restricao.")
            else:
                print("  AVISO: nao foi possivel limitar barcos distantes; relaxando a restricao.")
            local_best_routes, local_best_m9_remaining, local_top_scores = run_optimizer(
                enforce_all, False, require_zero_m9
            )

        return local_best_routes, local_best_m9_remaining, local_top_scores

    # Regra hard: se existir solucao com m9_restante=0, ela e obrigatoria.
    strict_routes, strict_m9_remaining, strict_top_scores = run_with_relaxations(require_zero_m9=True)
    if strict_routes:
        best_routes, best_m9_remaining, top_scores = strict_routes, strict_m9_remaining, strict_top_scores
    else:
        best_routes, best_m9_remaining, top_scores = run_with_relaxations(require_zero_m9=False)

    if top_scores:
        print("  Top combinacoes (m9_restante | score | dist | penalty | priority | comfort | pax_arrival | cluster):")
        for i, (m9_restante, score, dist, penalty, prio, comfort, pax_arrival, cluster, assignment) in enumerate(top_scores, 1):
            boat_pkgs = {b: [] for b in range(n_boats)}
            for pkg_idx, boat_idx in enumerate(assignment):
                boat_pkgs[boat_idx].append(pkg_idx + 1)
            parts = []
            for b_idx in range(n_boats):
                boat_name = boats[b_idx].name
                pkgs = ",".join(str(p) for p in boat_pkgs[b_idx]) if boat_pkgs[b_idx] else "-"
                parts.append(f"{boat_name}:{pkgs}")
            assign_str = " | ".join(parts)
            print(
                f"    {i}. {m9_restante} | {score:.2f} | {dist:.2f} | {penalty:.2f} | "
                f"{prio:.2f} | {comfort:.2f} | {pax_arrival:.2f} | {cluster:.2f}"
            )
            print(f"       {assign_str}")

    return best_routes or [], best_m9_remaining


def build_aqua_direct_route(boat: Boat, demands: List[Demand],
                             distances: Dict,
                             gangway_platforms: Set[str]) -> Optional[Route]:
    """
    Constrói rota direta para Aqua Helix (sem parada M9).
    Só leva pax TMIB para plataformas com gangway.
    A demanda M9 dessas plataformas fica para surfers via hub.
    """
    gangway_tmib = [d for d in demands
                    if d.platform_norm in gangway_platforms and d.tmib > 0]
    if not gangway_tmib:
        return None

    cap = boat.max_capacity

    # Ordenar por TMIB demand (maior primeiro)
    gangway_tmib.sort(key=lambda d: -d.tmib)

    stops = []
    total = 0
    priority_map = {d.platform_norm: d.priority for d in gangway_tmib}

    for d in gangway_tmib:
        if total + d.tmib <= cap:
            stops.append((d.platform_norm, d.tmib, 0))  # Só TMIB, sem M9
            total += d.tmib

    if not stops or total < 10:  # Mínimo para justificar uso do Aqua
        return None

    # Ordenação com prioridade a partir de TMIB
    if len(stops) > 1:
        stops = order_stops_with_priority(stops, distances, "TMIB", boat, priority_map)

    route = Route(
        boat=boat,
        stops=stops,
        m9_pickup=0,
        tmib_to_m9=0,
        uses_m9_hub=False,
        priority_map=priority_map,
    )
    route.total_distance = calc_route_distance(route, distances)
    return route


# ======== Main solver ========

def solve(config: Config, boats: List[Boat], demands: List[Demand],
          distances: Dict, gangway_platforms: Set[str]):
    warnings = []

    # ── Phase 1: Separar demanda M9 e processar rotas fixas ──
    m9_tmib_demand = 0
    m9_tmib_priority = 99
    platform_demands = []
    for d in demands:
        if short_plat(d.platform_norm) == "M9":
            m9_tmib_demand = d.tmib
            m9_tmib_priority = d.priority
        else:
            if d.tmib > 0 or d.m9 > 0:
                platform_demands.append(d.copy())

    available = [b for b in boats if b.available]
    if not available:
        print("ERRO: Nenhum barco disponivel.")
        return [], [], {}

    fixed_boats = [b for b in available if b.fixed_route]
    free_boats = [b for b in available if not b.fixed_route]

    results = []
    for boat in fixed_boats:
        results.append((boat, boat.fixed_route))
        deliveries = parse_fixed_route(boat.fixed_route)

        for plat_norm, delivered in deliveries.items():
            if short_plat(plat_norm) == "M9":
                m9_tmib_demand = max(0, m9_tmib_demand - delivered['tmib'])
                continue

            for d in platform_demands:
                if d.platform_norm == plat_norm:
                    d.tmib = max(0, d.tmib - delivered['tmib'])
                    d.m9 = max(0, d.m9 - delivered['m9'])
                    break

        print(f"  Rota fixa {boat.name}: subtraido da demanda")

    platform_demands = [d for d in platform_demands if d.total() > 0]

    # Classificar barcos
    surfers = [b for b in free_boats if not is_aqua_helix(b.name)]
    aquas = [b for b in free_boats if is_aqua_helix(b.name)]
    surfers.sort(key=lambda b: b.departure_minutes())
    aquas.sort(key=lambda b: b.departure_minutes())

    assigned_routes = []
    remaining_demands = list(platform_demands)
    remaining_m9_tmib = m9_tmib_demand

    # ── Phase 2: AQUA direct routes (prioridade) ──
    # Aqua Helix é melhor para rotas diretas de alta capacidade (sem M9 hub).
    # Só leva TMIB pax; M9 demand fica para surfers.
    for aqua in aquas[:]:
        route = build_aqua_direct_route(aqua, remaining_demands, distances,
                                         gangway_platforms)
        if route and route.total_pax() > 0:
            assigned_routes.append(route)
            aquas.remove(aqua)

            for stop in route.pre_m9_stops + route.stops:
                for d in remaining_demands[:]:
                    if d.platform_norm == stop[0]:
                        d.tmib = max(0, d.tmib - stop[1])
                        if d.total() <= 0:
                            remaining_demands.remove(d)
                        break

            print(f"  AQUA {aqua.name}: rota direta ({route.total_pax()} pax)")

    # ── Phase 3: Distant cluster dedication (PDO/PGA/PRB) ──
    if ENABLE_DISTANT_CLUSTER_DEDICATION:
        distant_demands = [d for d in remaining_demands
                           if get_geo_cluster(d.platform_norm) in ["PDO", "PGA", "PRB"]]

        if distant_demands:
            for boat in surfers[:]:
                route = build_m9_hub_route(boat, distant_demands, remaining_m9_tmib,
                                           distances, gangway_platforms, target_cluster=None)
                if route and route.total_pax() > 0:
                    assigned_routes.append(route)
                    surfers.remove(boat)

                    remaining_m9_tmib = max(0, remaining_m9_tmib - route.tmib_to_m9)

                    for stop in route.pre_m9_stops + route.stops:
                        for d in remaining_demands[:]:
                            if d.platform_norm == stop[0]:
                                d.tmib = max(0, d.tmib - stop[1])
                                d.m9 = max(0, d.m9 - stop[2])
                                if d.total() <= 0:
                                    remaining_demands.remove(d)
                                break
                    break

    def route_str_has_distant(route_str: str) -> bool:
        deliveries = parse_fixed_route(route_str)
        for plat_norm in deliveries.keys():
            if is_distant_cluster(get_geo_cluster(plat_norm)):
                return True
        return False

    def route_obj_has_distant(route: Route) -> bool:
        for stop in route.pre_m9_stops + route.stops:
            if is_distant_cluster(get_geo_cluster(stop[0])):
                return True
        return False

    distant_boats_already = 0
    for boat in fixed_boats:
        if boat.fixed_route and route_str_has_distant(boat.fixed_route):
            distant_boats_already += 1

    for route in assigned_routes:
        if route_obj_has_distant(route):
            distant_boats_already += 1

    # ── Phase 4: Combinatorial optimization ──
    # Agrupa demandas em pacotes (pares obrigatórios + individuais)
    # e testa todas as atribuições possíveis para minimizar distância total.
    remaining_boats = surfers + aquas
    remaining_demands = [d for d in remaining_demands if d.total() > 0]

    if remaining_demands and remaining_boats:
        packages = form_demand_packages(remaining_demands, remaining_boats)

        print(f"  Otimizador: {len(packages)} pacotes, {len(remaining_boats)} barcos, "
              f"{len(remaining_boats) ** len(packages)} combinacoes")

        hub_routes, remaining_m9_tmib = optimize_hub_assignments(
            packages, remaining_boats, distances, remaining_m9_tmib, gangway_platforms,
            m9_priority=m9_tmib_priority,
            distant_boats_already=distant_boats_already, max_distant_boats=1
        )
        assigned_routes.extend(hub_routes)

        # Atualizar remaining_demands
        for route in hub_routes:
            for stop in route.pre_m9_stops + route.stops:
                for d in remaining_demands[:]:
                    if d.platform_norm == stop[0]:
                        d.tmib = max(0, d.tmib - stop[1])
                        d.m9 = max(0, d.m9 - stop[2])
                        if d.total() <= 0:
                            remaining_demands.remove(d)
                        break

            # Remover barco da lista
            if route.boat in surfers:
                surfers.remove(route.boat)
            elif route.boat in aquas:
                aquas.remove(route.boat)

    # ── Phase 5: Fit remaining ──
    # Encaixar demandas restantes em rotas existentes com espaço
    remaining_demands = [d for d in remaining_demands if d.total() > 0]
    if remaining_demands:
        route_space = [(r, r.boat.max_capacity - r.max_load()) for r in assigned_routes]
        route_space.sort(key=lambda x: -x[1])

        for route, space in route_space:
            if not remaining_demands or space <= 0:
                continue

            is_aqua = is_aqua_helix(route.boat.name)
            route_clusters = set()
            for stop in route.stops:
                route_clusters.add(get_geo_cluster(stop[0]))

            remaining_demands.sort(key=lambda d: (d.priority if d.priority else 99, -d.total()))

            for d in remaining_demands[:]:
                if is_aqua and d.platform_norm not in gangway_platforms:
                    continue

                d_cluster = get_geo_cluster(d.platform_norm)
                compatible = not route_clusters or d_cluster in route_clusters
                if not compatible:
                    compatible = any(are_clusters_compatible(rc, d_cluster)
                                     for rc in route_clusters)

                if not compatible:
                    continue
                # Tentar adicionar a demanda e reequilibrar pre/post-M9
                temp_route = deepcopy(route)
                temp_route.stops.append((d.platform_norm, d.tmib, d.m9))
                if d.m9 > 0:
                    temp_route.m9_pickup += d.m9
                    temp_route.uses_m9_hub = True

                if rebuild_pre_m9(temp_route, distances):
                    route.stops = temp_route.stops
                    route.pre_m9_stops = temp_route.pre_m9_stops
                    route.m9_pickup = temp_route.m9_pickup
                    route.uses_m9_hub = temp_route.uses_m9_hub
                    route.priority_map[ d.platform_norm ] = d.priority
                    space = route.boat.max_capacity - route.max_load()
                    remaining_demands.remove(d)

    # ── Phase 6: Reordenar stops por NN e construir strings ──
    for route in assigned_routes:
        if route.uses_m9_hub and len(route.pre_m9_stops) > 1:
            route.pre_m9_stops = order_stops_with_priority(
                route.pre_m9_stops,
                distances,
                "TMIB",
                route.boat,
                route.priority_map,
            )

        if len(route.stops) > 1:
            start = norm_plat("M9") if route.uses_m9_hub else "TMIB"
            route.stops = order_stops_with_priority(
                route.stops,
                distances,
                start,
                route.boat,
                route.priority_map,
            )
        route.total_distance = calc_route_distance(route, distances)

    for route in assigned_routes:
        route_str = build_route_string(route)
        results.append((route.boat, route_str))

    # ── Phase 7: Verificações ──
    remaining_demands = [d for d in remaining_demands if d.total() > 0]
    if remaining_demands:
        warnings.append("\nDEMANDA NAO ATENDIDA:")
        for d in remaining_demands:
            warnings.append(f"  {d.platform}: TMIB={d.tmib}, M9={d.m9}")

    if remaining_m9_tmib > 0:
        warnings.append(f"\n{remaining_m9_tmib} pax TMIB->M9 nao alocados")

    total_tmib = sum(r.total_tmib() for r in assigned_routes)
    total_m9 = sum(r.total_m9() for r in assigned_routes)

    for boat in fixed_boats:
        deliveries = parse_fixed_route(boat.fixed_route)
        for plat, del_dict in deliveries.items():
            total_tmib += del_dict['tmib']
            total_m9 += del_dict['m9']

    summary = {
        'tmib_served': total_tmib,
        'm9_served': total_m9,
        'boats_used': len(results),
    }

    total_dist = sum(r.total_distance for r in assigned_routes)
    warnings.append(f"\nDistancia total (rotas livres): {total_dist:.1f} NM")

    return results, warnings, summary


def build_route_string(route: Route) -> str:
    """Constrói string de rota no formato padrão."""
    parts = []

    # TMIB
    total_tmib = route.total_tmib()
    if total_tmib > 0:
        parts.append(f"TMIB +{total_tmib}")
    else:
        parts.append("TMIB")

    # Paradas pré-M9
    if route.uses_m9_hub and route.pre_m9_stops:
        for stop in route.pre_m9_stops:
            plat_name = short_plat(stop[0])
            ops = []
            if stop[1] > 0:
                ops.append(f"-{stop[1]}")
            part = plat_name
            if ops:
                part += " " + " ".join(ops)
            parts.append(part)

    # M9 (se usa hub)
    if route.uses_m9_hub:
        m9_ops = []
        if route.tmib_to_m9 > 0:
            m9_ops.append(f"-{route.tmib_to_m9}")
        if route.m9_pickup > 0:
            m9_ops.append(f"+{route.m9_pickup}")
        m9_part = "M9"
        if m9_ops:
            m9_part += " " + " ".join(m9_ops)
        parts.append(m9_part)

    # Plataformas pós-M9
    for stop in route.stops:
        plat_name = short_plat(stop[0])
        ops = []
        if stop[1] > 0:  # TMIB drop
            ops.append(f"-{stop[1]}")
        if stop[2] > 0:  # M9 drop
            ops.append(f"(-{stop[2]})")

        part = plat_name
        if ops:
            part += " " + " ".join(ops)
        parts.append(part)

    return "/".join(parts)


# ======== Output ========

def write_output(results, warnings, summary, config, path):
    # Ordenar por horário de saída
    results.sort(key=lambda x: x[0].departure_minutes())

    lines = []
    for boat, route in results:
        line = f"{boat.name}  {boat.departure}  {route}"
        lines.append(line)

    with open(path, "w", encoding="utf-8") as f:
        f.write("DISTRIBUICAO DE PAX\n")
        f.write("=" * 70 + "\n")
        if config.troca_turma:
            f.write(f"Troca de turma: SIM | Rendidos em M9: {config.rendidos_m9}\n")
        f.write("\n")

        for line in lines:
            f.write(line + "\n")

        f.write("\n" + "-" * 70 + "\n")
        f.write(f"Resumo: {summary['tmib_served']} pax TMIB + {summary['m9_served']} pax M9 = ")
        f.write(f"{summary['tmib_served'] + summary['m9_served']} pax total\n")
        f.write(f"Barcos utilizados: {summary['boats_used']}\n")
        f.write("=" * 70 + "\n")

        if warnings:
            f.write("\n")
            for w in warnings:
                f.write(w + "\n")

    print(f"\nDistribuicao salva em: {path}\n")
    for line in lines:
        print(f"  {line}")

    print(f"\n  Resumo: {summary['tmib_served']} TMIB + {summary['m9_served']} M9 = ", end="")
    print(f"{summary['tmib_served'] + summary['m9_served']} pax | {summary['boats_used']} barcos")

    if warnings:
        print()
        for w in warnings:
            print(f"  {w}")


# ======== Main ========

def main():
    if not os.path.exists(INPUT_FILE):
        print(f"ERRO: '{INPUT_FILE}' nao encontrado. Execute criarInputSolver.py primeiro.")
        return

    if not os.path.exists(DIST_FILE):
        print(f"ERRO: '{DIST_FILE}' nao encontrado.")
        return

    distances = load_distances(DIST_FILE)
    speeds = load_speeds(SPEED_FILE)
    gangway_platforms = load_gangway(GANGWAY_FILE)
    config, boats, demands = read_solver_input(INPUT_FILE)

    if gangway_platforms:
        print(f"Plataformas com gangway: {len(gangway_platforms)}")
    else:
        print("AVISO: gangway.json nao encontrado - Aqua Helix nao podera operar")

    for boat in boats:
        boat.speed = get_speed(speeds, boat.name)

    n_available = sum(1 for b in boats if b.available)
    n_demand = sum(1 for d in demands if d.tmib > 0 or d.m9 > 0)
    total_tmib = sum(d.tmib for d in demands)
    total_m9 = sum(d.m9 for d in demands if short_plat(d.platform_norm) != "M9")

    print("SOLVER DE DISTRIBUICAO DE PAX v4")
    print("=" * 40)
    print(f"Troca de turma: {'SIM' if config.troca_turma else 'NAO'}")
    print(f"Barcos disponiveis: {n_available}")
    print(f"Plataformas com demanda: {n_demand}")
    print(f"Total pax: {total_tmib} TMIB + {total_m9} M9 = {total_tmib + total_m9}")
    print("=" * 40)

    if n_available == 0:
        print("\nERRO: Nenhum barco disponivel. Marque 'SIM' na coluna Disponivel.")
        return

    if n_demand == 0:
        print("\nERRO: Nenhuma demanda informada. Preencha a tabela de demanda.")
        return

    results, warnings, summary = solve(config, boats, demands, distances, gangway_platforms)

    if results:
        write_output(results, warnings, summary, config, OUTPUT_FILE)


if __name__ == "__main__":
    main()
