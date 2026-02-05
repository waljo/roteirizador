#!/usr/bin/env python3
"""
Importa uma distribuicao (rotas compactas) de um arquivo .txt
e grava a demanda agregada em solver_input.xlsx.

Uso:
  python importar_distribuicao.py --input distribuicao.txt --output solver_input.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
from typing import Dict, Tuple

from openpyxl import load_workbook

import solver
from criarInputSolver import PLATFORMS


RE_TIME = re.compile(r"\\b\\d{2}:\\d{2}\\b")


def _extract_route(line: str) -> str:
    if "TMIB" in line:
        idx = line.index("TMIB")
        return line[idx:].strip()
    m = RE_TIME.search(line)
    if m:
        return line[m.end():].strip()
    return ""


def _parse_distribution_file(path: str) -> Dict[str, Dict[str, int]]:
    deliveries: Dict[str, Dict[str, int]] = {}
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            up = line.upper()
            if up.startswith("DISTRIBUICAO") or up.startswith("=") or up.startswith("-"):
                continue
            if "TMIB" not in up:
                continue
            route_str = _extract_route(line)
            if not route_str:
                continue
            route_deliveries = solver.parse_fixed_route(route_str)
            for plat_norm, delivered in route_deliveries.items():
                if plat_norm not in deliveries:
                    deliveries[plat_norm] = {"tmib": 0, "m9": 0}
                deliveries[plat_norm]["tmib"] += delivered.get("tmib", 0)
                deliveries[plat_norm]["m9"] += delivered.get("m9", 0)
    return deliveries


def _find_demand_header_row(ws) -> int:
    for r in range(1, ws.max_row + 1):
        if (
            ws.cell(row=r, column=2).value == "Plataforma"
            and ws.cell(row=r, column=3).value == "M9"
            and ws.cell(row=r, column=4).value == "TMIB"
        ):
            return r
    raise ValueError("Cabecalho de demanda nao encontrado.")


def _find_demand_end_row(ws, start_row: int) -> int:
    for r in range(start_row, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if isinstance(val, str) and val.strip().upper().startswith("SINTAXE DAS ROTAS"):
            return r - 1
    return ws.max_row


def _clear_demand_table(ws, start_row: int, end_row: int):
    for r in range(start_row, end_row + 1):
        for c in range(2, 6):
            cell = ws.cell(row=r, column=c)
            cell.value = None


def _write_demands(ws, start_row: int, deliveries: Dict[str, Dict[str, int]]):
    demand_by_short: Dict[str, Tuple[int, int]] = {}
    for plat_norm, delivered in deliveries.items():
        short = solver.short_plat(plat_norm)
        m9 = delivered.get("m9", 0)
        tmib = delivered.get("tmib", 0)
        if short not in demand_by_short:
            demand_by_short[short] = (m9, tmib)
        else:
            prev_m9, prev_tmib = demand_by_short[short]
            demand_by_short[short] = (prev_m9 + m9, prev_tmib + tmib)

    row = start_row
    for plat in PLATFORMS:
        m9, tmib = demand_by_short.get(plat, (0, 0))
        if m9 == 0 and tmib == 0:
            continue
        ws.cell(row=row, column=2, value=plat)
        ws.cell(row=row, column=3, value=m9)
        ws.cell(row=row, column=4, value=tmib)
        row += 1


def run_import(input_path: str, output_path: str):
    if not os.path.exists(input_path):
        raise SystemExit(f"Arquivo nao encontrado: {input_path}")
    if not os.path.exists(output_path):
        raise SystemExit(f"Arquivo nao encontrado: {output_path}")

    deliveries = _parse_distribution_file(input_path)

    wb = load_workbook(output_path)
    ws = wb.active

    header_row = _find_demand_header_row(ws)
    demand_start = header_row + 1
    demand_end = _find_demand_end_row(ws, demand_start)

    _clear_demand_table(ws, demand_start, demand_end)
    _write_demands(ws, demand_start, deliveries)

    wb.save(output_path)
    print(f"Demanda gravada em {output_path} a partir de {input_path}")


def main():
    parser = argparse.ArgumentParser(description="Importar distribuicao para solver_input.xlsx")
    parser.add_argument("--input", default="distribuicao.txt", help="Arquivo de distribuicao .txt")
    parser.add_argument("--output", default="solver_input.xlsx", help="Arquivo solver_input.xlsx")
    args = parser.parse_args()

    run_import(args.input, args.output)


if __name__ == "__main__":
    main()
