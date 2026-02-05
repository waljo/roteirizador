# criarInputSolver.py
# Cria planilha de entrada para o solver de distribuição de PAX

import os
import re
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


SPEED_FILE = "velocidades.txt"
OUTPUT_FILE = "solver_input.xlsx"
DEFAULT_SPEED_KN = 14.0

DEFAULT_BOATS = [
    "SURFER 1870",
    "SURFER 1871",
    "SURFER 1906",
    "SURFER 1930",
    "SURFER 1931",
    "AQUA HELIX",
]

PLATFORMS = [
    "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11",
    "B1", "B2", "B3", "B4",
    "PGA1", "PGA2", "PGA3", "PGA4", "PGA5", "PGA7", "PGA8",
    "PDO1", "PDO2", "PDO3",
    "PRB1",
]


def load_speeds(path: str) -> Dict[str, float]:
    speeds: Dict[str, float] = {}
    if not os.path.exists(path):
        return speeds
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    current_section = None
    for line in content.split("\n"):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        section_match = re.match(r"\[([A-Z_]+)\]", line)
        if section_match:
            current_section = section_match.group(1).replace("_", " ")
            continue
        if "=" in line:
            parts = line.split("=")
            if len(parts) == 2:
                left_part = parts[0].strip()
                try:
                    speed = float(parts[1].strip())
                except ValueError:
                    continue
                if current_section and left_part.isdigit():
                    full_name = f"{current_section} {left_part}"
                    speeds[full_name.upper()] = speed
                else:
                    name = left_part.replace("_", " ").upper()
                    speeds[name] = speed
                    speeds[left_part.upper()] = speed
    return speeds


def create_solver_input():
    speeds = load_speeds(SPEED_FILE)
    if speeds:
        boat_names = sorted(set(speeds.keys()))
    else:
        boat_names = DEFAULT_BOATS

    wb = Workbook()
    ws = wb.active
    ws.title = "Solver Input"
    ws.sheet_view.showGridLines = False

    # Hidden sheet with platform list for dropdowns
    list_ws = wb.create_sheet("Lists")
    for i, plat in enumerate(PLATFORMS, start=1):
        list_ws.cell(row=i, column=1, value=plat)
    list_ws.sheet_state = "hidden"

    # Estilos
    thin = Side(style="thin", color="4F4F4F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    blue_fill = PatternFill("solid", fgColor="BDD7EE")
    light_header_fill = PatternFill("solid", fgColor="D9E2F3")
    section_font = Font(bold=True, size=11)
    label_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # ── TITULO ──
    ws.merge_cells("B1:H1")
    title = ws.cell(row=1, column=2, value="SOLVER DE DISTRIBUIÇÃO DE PAX")
    title.font = Font(bold=True, size=14)
    title.alignment = center
    ws.row_dimensions[1].height = 30

    # ── CONFIGURAÇÃO ──
    r = 3
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2, value="CONFIGURAÇÃO").font = section_font
    for col in range(2, 6):
        ws.cell(row=r, column=col).fill = blue_fill
        ws.cell(row=r, column=col).border = border

    r = 4
    ws.cell(row=r, column=2, value="Troca de turma?").font = label_font
    troca_cell = ws.cell(row=r, column=3, value="NÃO")
    troca_cell.border = border
    troca_cell.alignment = center
    dv_troca = DataValidation(type="list", formula1='"SIM,NÃO"', allow_blank=False)
    ws.add_data_validation(dv_troca)
    dv_troca.add(troca_cell)

    r = 5
    ws.cell(row=r, column=2, value="Rendidos em M9:").font = label_font
    rendidos_cell = ws.cell(row=r, column=3, value=0)
    rendidos_cell.border = border
    rendidos_cell.alignment = center

    # ── BARCOS DISPONÍVEIS ──
    r = 7
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(row=r, column=2, value="BARCOS DISPONÍVEIS").font = section_font
    for col in range(2, 9):
        ws.cell(row=r, column=col).fill = blue_fill
        ws.cell(row=r, column=col).border = border

    r = 8
    for col, name in [(2, "Barco"), (3, "Disponível"), (4, "Hora Saída")]:
        cell = ws.cell(row=r, column=col, value=name)
        cell.font = label_font
        cell.fill = light_header_fill
        cell.alignment = center
        cell.border = border
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    rf_header = ws.cell(row=r, column=5, value="Rota Fixa (opcional)")
    rf_header.font = label_font
    rf_header.fill = light_header_fill
    rf_header.alignment = center
    for col in range(5, 9):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).fill = light_header_fill

    dv_disp = DataValidation(type="list", formula1='"SIM,NÃO"', allow_blank=False)
    ws.add_data_validation(dv_disp)

    boat_start = 9
    for i, boat in enumerate(boat_names):
        r = boat_start + i
        ws.cell(row=r, column=2, value=boat).border = border
        disp = ws.cell(row=r, column=3, value="NÃO")
        disp.border = border
        disp.alignment = center
        dv_disp.add(disp)
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=4).alignment = center
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=5).alignment = left
        for col in range(6, 9):
            ws.cell(row=r, column=col).border = border

    boat_end = boat_start + len(boat_names) - 1

    # ── DEMANDA DE DISTRIBUIÇÃO ──
    r = boat_end + 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2, value="DEMANDA DE DISTRIBUIÇÃO").font = section_font
    for col in range(2, 6):
        ws.cell(row=r, column=col).fill = blue_fill
        ws.cell(row=r, column=col).border = border

    r += 1
    for col, name in [(2, "Plataforma"), (3, "M9"), (4, "TMIB"), (5, "Prioridade")]:
        cell = ws.cell(row=r, column=col, value=name)
        cell.font = label_font
        cell.fill = light_header_fill
        cell.alignment = center
        cell.border = border

    dv_prio = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    ws.add_data_validation(dv_prio)

    demand_start = r + 1
    for i, plat in enumerate(PLATFORMS):
        r = demand_start + i
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=3).alignment = center
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=4).alignment = center
        prio = ws.cell(row=r, column=5)
        prio.border = border
        prio.alignment = center
        dv_prio.add(prio)

    # Plataforma dropdown list (instead of fixed names)
    platform_range = f"=Lists!$A$1:$A${len(PLATFORMS)}"
    dv_platform = DataValidation(type="list", formula1=platform_range, allow_blank=True)
    ws.add_data_validation(dv_platform)
    for i in range(len(PLATFORMS)):
        r = demand_start + i
        dv_platform.add(ws.cell(row=r, column=2))

    demand_end = demand_start + len(PLATFORMS) - 1

    # ── SINTAXE DAS ROTAS ──
    r = demand_end + 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(row=r, column=2, value="SINTAXE DAS ROTAS (para rotas fixas)").font = section_font
    for col in range(2, 9):
        ws.cell(row=r, column=col).fill = blue_fill
        ws.cell(row=r, column=col).border = border

    syntax_lines = [
        ("+N = embarque no TMIB  |  -N = desembarque de pax do TMIB", False),
        ("+N (em M9) = embarque em M9  |  (-N) = desembarque de pax de M9", False),
        ("{DEST:+N} = embarque de transbordo com destino DEST", False),
        ("{ORIG:-N} = desembarque de transbordo vindo de ORIG", False),
        ("Separar paradas com /", False),
        ("", False),
        ("Exemplo: TMIB +22/M9 -6 +6/M6 -2 {B1:+1}/B2 -14/B1 (-6) {M6:-1}", True),
        ("  → Sai de TMIB com 22 pax", False),
        ("  → Em M9: desembarca 6 TMIB, embarca 6 M9", False),
        ("  → Em M6: desembarca 2 TMIB, pega 1 transbordo para B1", False),
        ("  → Em B2: desembarca 14 TMIB", False),
        ("  → Em B1: desembarca 6 M9, deixa 1 transbordo de M6", False),
    ]

    for text, bold in syntax_lines:
        r += 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=2, value=text)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if bold:
            cell.font = Font(bold=True)

    # ── Larguras das colunas ──
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    wb.save(OUTPUT_FILE)
    print(f"Planilha criada: {OUTPUT_FILE}")


if __name__ == "__main__":
    create_solver_input()
