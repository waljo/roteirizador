# criarTabela6.py
# Versao com entrada via Excel ao inves de viagens.txt
# Gera uma planilha no layout "antigo" da PROGRAMACAO DE DISTRIBUICAO DE PAX

import json
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont


# =========================
# CONFIG
# =========================
DIST_FILE = "distplat.json"
SPEED_FILE = "velocidades.txt"
INPUT_FILE = "viagens_input.xlsx"
OUT_FILE = "programacao_pax.xlsx"

MINUTES_PER_PAX = 1
DEFAULT_SPEED_KN = 14.0
MAX_CAPACITY_SURFER = 24
MAX_CAPACITY_AQUA_HELIX = 100

SHEET_NAME = "DISTRIBUICAO - CL"
TITLE_TEXT = "PROGRAMACAO DE DISTRIBUICAO DE PAX"

COLS = [
    "EMBARCACAO",
    "HR SAIDA ORIGEM",
    "QUANT PAX EMBARQUE",
    "ORIGEM",
    "HORA CHEGADA DESTINO",
    "DESTINO",
    "QUANT. PAX DESEMBARQUE",
    "QUANT PAX A BORDO",
    "LINGADAS",
    "OPERACAO DE TRANSBORDO",
    "OPERACAO",
]

# Lista de plataformas
PLATFORMS = [
    "TMIB",
    "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11",
    "B1", "B2", "B3", "B4",
    "PGA1", "PGA2", "PGA3", "PGA4", "PGA5", "PGA7", "PGA8",
    "PDO1", "PDO2", "PDO3",
    "PRB1",
]


# =========================
# ALIASES
# =========================
def norm_plat(code: str) -> str:
    c = code.strip().upper()
    if re.match(r"^(TMIB|NORWIND GALE)$", c):
        return c
    if re.match(r"^(PCM|PCB|PGA|PRB|PDO)-\d{2}$", c):
        return c
    m = re.match(r"^M(\d+)$", c)
    if m:
        return f"PCM-{int(m.group(1)):02d}"
    b = re.match(r"^B(\d+)$", c)
    if b:
        return f"PCB-{int(b.group(1)):02d}"
    pg = re.match(r"^PGA(\d+)$", c)
    if pg:
        return f"PGA-{int(pg.group(1)):02d}"
    pdo = re.match(r"^PDO(\d+)$", c)
    if pdo:
        return f"PDO-{int(pdo.group(1)):02d}"
    prb = re.match(r"^PRB(\d+)$", c)
    if prb:
        return f"PRB-{int(prb.group(1)):02d}"
    m2 = re.match(r"^PCM(\d+)$", c)
    if m2:
        return f"PCM-{int(m2.group(1)):02d}"
    b2 = re.match(r"^PCB(\d+)$", c)
    if b2:
        return f"PCB-{int(b2.group(1)):02d}"
    c = re.sub(r"\s*\(.*?\)\s*$", "", c).strip()
    return c


def short_plat(norm: str) -> str:
    n = norm.upper().strip()
    if n == "TMIB":
        return "TMIB"
    if re.match(r"^PCM-\d{2}$", n):
        return f"M{int(n.split('-')[1])}"
    if re.match(r"^PCB-\d{2}$", n):
        return f"B{int(n.split('-')[1])}"
    if re.match(r"^PGA-\d{2}$", n):
        return f"PGA{int(n.split('-')[1])}"
    if re.match(r"^PDO-\d{2}$", n):
        return f"PDO{int(n.split('-')[1])}"
    if re.match(r"^PRB-\d{2}$", n):
        return f"PRB{int(n.split('-')[1])}"
    return n


# =========================
# PARSE: velocidades.txt
# =========================
def load_speeds(path: str) -> Dict[str, float]:
    speeds: Dict[str, float] = {}
    if not os.path.exists(path):
        return speeds

    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    current_section = None
    for line in content.split("\n"):
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        section_match = re.match(r"\[([A-Z_]+)\]", line)
        if section_match:
            current_section = section_match.group(1).replace("_", " ")
            continue

        if "=" in line:
            parts = line.split("=")
            if len(parts) == 2:
                left_part = parts[0].strip()
                try:
                    speed = float(parts[1].strip())
                except ValueError:
                    continue

                if current_section and left_part.isdigit():
                    full_name = f"{current_section} {left_part}"
                    speeds[full_name.upper()] = speed
                else:
                    name = left_part.replace("_", " ").upper()
                    speeds[name] = speed
                    speeds[left_part.upper()] = speed

    return speeds


# =========================
# PARSE: distplat.json
# =========================
def load_distances_json(path: str) -> Dict[str, Dict[str, float]]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Arquivo de distancias nao encontrado: {path}")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    dist: Dict[str, Dict[str, float]] = {}
    for a, m in data.items():
        aN = norm_plat(str(a))
        dist.setdefault(aN, {})
        for b, v in m.items():
            bN = norm_plat(str(b))
            dist[aN][bN] = float(v)
    return dist


def get_distance_nm(dist: Dict[str, Dict[str, float]], a: str, b: str) -> float:
    aN, bN = norm_plat(a), norm_plat(b)
    if aN in dist and bN in dist[aN]:
        return dist[aN][bN]
    if bN in dist and aN in dist[bN]:
        return dist[bN][aN]
    raise KeyError(f"Distancia nao encontrada entre {a} e {b}")


# =========================
# ROTA: modelo
# =========================
@dataclass
class StopOps:
    plat: str
    board_tmib: int = 0
    board_m9: int = 0
    disemb_tmib: int = 0
    disemb_m9: int = 0
    board_other: Dict[str, int] = None
    disemb_other: Dict[str, int] = None

    def __post_init__(self):
        if self.board_other is None:
            self.board_other = {}
        if self.disemb_other is None:
            self.disemb_other = {}


# =========================
# PRE-PROCESSAMENTO: M9 pickups por ocorrencia
# =========================
def preprocess_m9_pickups_by_index(stops: List[StopOps]) -> Dict[int, Dict[str, int]]:
    out: Dict[int, Dict[str, int]] = {}

    for i, s in enumerate(stops):
        if short_plat(norm_plat(s.plat)) != "M9":
            continue
        if s.board_m9 <= 0:
            continue

        remaining = s.board_m9
        out[i] = {}

        for k in range(i + 1, len(stops)):
            if remaining <= 0:
                break
            sk = stops[k]
            if sk.disemb_m9 > 0:
                take = min(sk.disemb_m9, remaining)
                destN = norm_plat(sk.plat)
                out[i][destN] = out[i].get(destN, 0) + take
                remaining -= take

    return out


# =========================
# SIMULACAO
# =========================
def hhmm_to_dt(hhmm: str) -> datetime:
    return datetime(2000, 1, 1, int(hhmm[:2]), int(hhmm[3:5]))


def dt_to_hhmm(dt: datetime) -> str:
    return dt.strftime("%H:%M")


def travel_minutes(distance_nm: float, speed_kn: float) -> int:
    mins = distance_nm / speed_kn * 60.0
    return int(mins + 0.999999)


def make_summary_path(stops: List[StopOps]) -> str:
    return " > ".join(short_plat(norm_plat(s.plat)) for s in stops)


def fmt_desemb(tmib_n: int, m9_n: int, other: Dict[str, int]) -> str:
    base = ""
    if tmib_n > 0:
        base += f"{tmib_n}"
    else:
        base += "0"

    if m9_n > 0:
        base += f"({m9_n})"

    if other:
        parts = []
        for k, v in sorted(other.items()):
            parts.append(f"{short_plat(norm_plat(k))}:{v}")
        base += " {" + ", ".join(parts) + "}"

    if (tmib_n == 0) and (m9_n == 0) and (not other):
        return "0"
    return base


def build_pick_comment_for_stop(
    stop_index: int,
    stop: StopOps,
    m9_pickups_by_index: Dict[int, Dict[str, int]],
) -> str:
    pshort = short_plat(norm_plat(stop.plat))

    if pshort == "M9" and stop.board_m9 > 0:
        dist_map = m9_pickups_by_index.get(stop_index, {})
        if dist_map:
            items = []
            for destN, n in sorted(dist_map.items(), key=lambda x: x[0]):
                items.append(f"{n} pax p/ {short_plat(destN)}")
            return "Em M9 pega " + " e ".join(items)
        return f"Em M9 pega {stop.board_m9} pax"

    if stop.board_other:
        items = []
        for dest_raw, n in sorted(stop.board_other.items(), key=lambda x: norm_plat(x[0])):
            items.append(f"{n} pax p/ {short_plat(norm_plat(dest_raw))}")
        return f"Em {pshort} pega " + " e ".join(items)

    return ""


def get_max_capacity(vessel_name: str) -> int:
    if "AQUA" in vessel_name.upper() and "HELIX" in vessel_name.upper():
        return MAX_CAPACITY_AQUA_HELIX
    return MAX_CAPACITY_SURFER


def simulate_trip(
    dist: Dict[str, Dict[str, float]],
    vessel_name: str,
    start_hhmm: str,
    stops: List[StopOps],
    speed_kn: float,
    minutes_per_pax: int = 1,
) -> Tuple[List[Dict], int, str, float]:
    if len(stops) < 2:
        raise ValueError("Rota precisa ter pelo menos 2 plataformas")

    max_capacity = get_max_capacity(vessel_name)
    m9_pickups_by_index = preprocess_m9_pickups_by_index(stops)

    onboard_tmib = 0
    onboard_m9 = 0
    onboard_other: Dict[str, int] = {}

    current_time = hhmm_to_dt(start_hhmm)

    first = stops[0]
    if short_plat(norm_plat(first.plat)) == "TMIB":
        onboard_tmib += first.board_tmib
    elif short_plat(norm_plat(first.plat)) == "M9":
        onboard_m9 += first.board_m9

    for dest, n in first.board_other.items():
        oN = norm_plat(first.plat)
        onboard_other[oN] = onboard_other.get(oN, 0) + n

    rows: List[Dict] = []
    total_pax_moved = 0

    for i in range(len(stops) - 1):
        orig = stops[i]
        dest = stops[i + 1]

        orig_short = short_plat(norm_plat(orig.plat))
        dest_short = short_plat(norm_plat(dest.plat))

        depart_time = current_time
        pax_embarked_this_leg = onboard_tmib + onboard_m9 + sum(onboard_other.values())

        d_nm = get_distance_nm(dist, orig.plat, dest.plat)
        travel_min = travel_minutes(d_nm, speed_kn)
        arrive_time = depart_time + timedelta(minutes=travel_min)

        dis_tmib = dest.disemb_tmib
        dis_m9 = dest.disemb_m9
        dis_other: Dict[str, int] = {}
        for k, v in dest.disemb_other.items():
            dis_other[k] = dis_other.get(k, 0) + v

        is_last_leg = (i == len(stops) - 2)
        if is_last_leg:
            if dis_tmib == 0:
                dis_tmib = onboard_tmib
            if dis_m9 == 0:
                dis_m9 = onboard_m9
            for oN, q in onboard_other.items():
                if q > 0:
                    dis_other[oN] = dis_other.get(oN, 0) + q

        if dis_tmib > onboard_tmib:
            dis_tmib = onboard_tmib
        onboard_tmib -= dis_tmib

        if dis_m9 > onboard_m9:
            dis_m9 = onboard_m9
        onboard_m9 -= dis_m9

        for o, q in list(dis_other.items()):
            oN = norm_plat(o)
            have = onboard_other.get(oN, 0)
            if q > have:
                q = have
                dis_other[o] = q
            onboard_other[oN] = have - q
            if onboard_other[oN] == 0:
                onboard_other.pop(oN, None)

        board_tmib = dest.board_tmib
        board_m9 = dest.board_m9
        board_other = dest.board_other

        if dest_short == "TMIB":
            onboard_tmib += board_tmib
        elif dest_short == "M9":
            onboard_m9 += board_m9

        for dest2, q in board_other.items():
            oN = norm_plat(dest.plat)
            onboard_other[oN] = onboard_other.get(oN, 0) + q

        moved_ops = (dis_tmib + dis_m9 + sum(dis_other.values())) + (board_tmib + board_m9 + sum(board_other.values()))
        total_pax_moved += (dis_tmib + dis_m9 + sum(dis_other.values()))

        current_time = arrive_time + timedelta(minutes=moved_ops * minutes_per_pax)

        pax_after_ops = onboard_tmib + onboard_m9 + sum(onboard_other.values())
        if pax_after_ops > max_capacity:
            raise ValueError(
                f"ERRO: Capacidade excedida na chegada em {dest_short}! "
                f"A bordo: {pax_after_ops} PAX (maximo permitido: {max_capacity})"
            )

        if dest_short == "TMIB":
            op_transb = "NAO"
        else:
            op_transb = "SIM" if (dis_m9 > 0 or any(v > 0 for v in dis_other.values())) else "NAO"

        obs = build_pick_comment_for_stop(i + 1, dest, m9_pickups_by_index)

        desemb_txt = fmt_desemb(dis_tmib, dis_m9, {k: v for k, v in dis_other.items() if v > 0})

        rows.append({
            "HR SAIDA ORIGEM": dt_to_hhmm(depart_time),
            "QUANT PAX EMBARQUE": pax_embarked_this_leg,
            "ORIGEM": orig_short,
            "HORA CHEGADA DESTINO": dt_to_hhmm(arrive_time),
            "DESTINO": dest_short,
            "QUANT. PAX DESEMBARQUE": desemb_txt,
            "QUANT PAX A BORDO": pax_after_ops,
            "LINGADAS": "",
            "OPERACAO DE TRANSBORDO": op_transb,
            "OPERACAO": obs,
        })

    summary_line = f"{make_summary_path(stops)} = {total_pax_moved} PAX TRANSPORTADOS"
    return rows, total_pax_moved, summary_line, speed_kn


# =========================
# PARSE ROUTE (from text syntax)
# =========================
BRACE_RE = re.compile(r"\{([^{}]+)\}")


def parse_stop_part(part: str) -> StopOps:
    p = part.strip()
    if not p:
        raise ValueError("Trecho vazio na rota")
    tokens = p.split()
    plat = tokens[0].strip()
    rest = p[len(tokens[0]):].strip()
    ops = StopOps(plat=plat)

    for m in BRACE_RE.finditer(rest):
        inside = m.group(1).strip()
        if ":" not in inside:
            raise ValueError(f"Chave invalida: {{{inside}}}")
        left, right = [x.strip() for x in inside.split(":", 1)]
        target = left
        val = right.replace(" ", "")
        if not re.match(r"^[+-]\d+$", val):
            raise ValueError(f"Valor invalido em chave: {{{inside}}}")
        n = int(val)
        if n > 0:
            ops.board_other[target] = ops.board_other.get(target, 0) + n
        else:
            ops.disemb_other[target] = ops.disemb_other.get(target, 0) + abs(n)

    rest_wo_braces = BRACE_RE.sub(" ", rest)

    for tok in rest_wo_braces.split():
        t = tok.strip()
        if not t:
            continue
        if re.match(r"^\+\d+$", t):
            n = int(t[1:])
            if plat.upper() == "TMIB":
                ops.board_tmib += n
            elif plat.upper() == "M9" or plat.upper() == "PCM-09":
                ops.board_m9 += n
            else:
                raise ValueError(f"Embarque '+{n}' em {plat} nao permitido sem destino")
        elif re.match(r"^\(-\d+\)$", t):
            n = int(t[2:-1])
            ops.disemb_m9 += abs(n)
        elif re.match(r"^-\d+$", t):
            n = int(t[1:])
            ops.disemb_tmib += n

    return ops


def parse_route(route: str) -> List[StopOps]:
    parts = [p.strip() for p in route.strip().split("/") if p.strip()]
    return [parse_stop_part(p) for p in parts]


# =========================
# CRIAR TEMPLATE DE ENTRADA
# =========================
def create_input_template(path: str, speeds: Dict[str, float]):
    """Cria o arquivo Excel de entrada com dropdowns e instrucoes."""
    wb = Workbook()

    # ========== Aba Viagens ==========
    ws_viagens = wb.active
    ws_viagens.title = "Viagens"

    # Row 1: Explanation text (merged A1:C1)
    explanation = (
        "COMO DIGITAR A ROTA: +x = embarque no TMIB | -x = desembarque de pax do TMIB | "
        "+x (em M9) = embarque em M9 | (-x) = desembarque de pax de M9 | "
        "{destino:+x} = embarque em outra plataforma (ex: {B1:+1}) | "
        "{origem:-x} = desembarque de pax de outra origem (ex: {M6:-1})\n"
        "EXEMPLO: Lancha pega 22 no TMIB → M10 deixa 5 → M9 deixa 7 e pega 4 p/ B2 → "
        "M6 deixa 4 e pega 1 p/ B1 → B2 deixa 8 (4 TMIB + 4 M9) → B1 deixa 3 (2 TMIB + 1 M6). "
        "Rota: TMIB +22/M10 -5/M9 -7 +4/M6 -4 {B1:+1}/B2 -4 (-4)/B1 -2 {M6:-1}"
    )
    ws_viagens.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    explanation_cell = ws_viagens.cell(row=1, column=1, value=explanation)
    explanation_cell.alignment = Alignment(wrap_text=True, vertical="center")
    explanation_cell.font = Font(size=10)
    ws_viagens.row_dimensions[1].height = 45

    # Row 2: Headers
    headers = ["Embarcacao", "Hora Saida", "Rota"]

    # Header style
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws_viagens.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    ws_viagens.column_dimensions["A"].width = 20
    ws_viagens.column_dimensions["B"].width = 12
    ws_viagens.column_dimensions["C"].width = 80

    ws_viagens.row_dimensions[2].height = 24

    # ========== Aba Embarcacoes ==========
    ws_vessels = wb.create_sheet("Embarcacoes")
    ws_vessels.cell(row=1, column=1, value="Embarcacao").font = Font(bold=True)
    ws_vessels.cell(row=1, column=2, value="Velocidade (nos)").font = Font(bold=True)

    # Extrair nomes unicos de embarcacoes do arquivo de velocidades
    vessel_names = sorted(set(speeds.keys()))
    if not vessel_names:
        vessel_names = ["SURFER 1870", "SURFER 1871", "SURFER 1905", "SURFER 1930", "SURFER 1931", "AQUA HELIX"]

    for i, name in enumerate(vessel_names, start=2):
        ws_vessels.cell(row=i, column=1, value=name)
        ws_vessels.cell(row=i, column=2, value=speeds.get(name, DEFAULT_SPEED_KN))

    ws_vessels.column_dimensions["A"].width = 20
    ws_vessels.column_dimensions["B"].width = 18

    # ========== Aba Instrucoes ==========
    ws_inst = wb.create_sheet("Instrucoes")
    instructions = [
        "INSTRUCOES PARA PREENCHIMENTO",
        "",
        "Cada linha representa uma VIAGEM completa.",
        "",
        "COLUNAS:",
        "  - Embarcacao: Nome da embarcacao (selecione da lista ou digite)",
        "  - Hora Saida: Horario de saida da origem no formato HH:MM",
        "  - Rota: Rota completa usando a sintaxe com '/'",
        "",
        "SINTAXE DA ROTA:",
        "  - Separar paradas com '/'",
        "  - +N = embarcar N passageiros (pool TMIB em TMIB, pool M9 em M9)",
        "  - -N = desembarcar N passageiros do pool TMIB",
        "  - (-N) = desembarcar N passageiros do pool M9",
        "  - {DESTINO:+N} = pegar N pax de transbordo com destino DESTINO",
        "  - {ORIGEM:-N} = deixar N pax de transbordo vindos de ORIGEM",
        "",
        "EXEMPLOS DE ROTAS:",
        "  TMIB +22/M9 -6 +6/M6 -2/B2 -14/B1 (-6)",
        "    -> Sai de TMIB com 22 pax",
        "    -> Em M9: desembarca 6 TMIB, embarca 6 M9",
        "    -> Em M6: desembarca 2 TMIB",
        "    -> Em B2: desembarca 14 TMIB",
        "    -> Em B1: desembarca 6 M9",
        "",
        "  M9 +22/M6 (-11) {TMIB:+1}/B1 (-5)/TMIB (-6) {M6:-1}",
        "    -> Sai de M9 com 22 pax (pool M9)",
        "    -> Em M6: desembarca 11 M9, pega 1 transbordo p/ TMIB",
        "    -> Em B1: desembarca 5 M9",
        "    -> Em TMIB: desembarca 6 M9, deixa 1 transbordo de M6",
        "",
        "PLATAFORMAS DISPONIVEIS:",
        "  TMIB, M1-M11, B1-B4, PGA1-PGA8, PDO1-PDO3, PRB1",
        "",
        "NOTAS:",
        "  - Capacidade maxima: 24 passageiros",
        "  - Tempos de chegada sao calculados automaticamente",
    ]

    for i, line in enumerate(instructions, start=1):
        cell = ws_inst.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)

    ws_inst.column_dimensions["A"].width = 80

    # ========== Data Validation (Dropdowns) ==========
    # Vessel dropdown
    vessel_dv = DataValidation(
        type="list",
        formula1="Embarcacoes!$A$2:$A$50",
        allow_blank=True
    )
    vessel_dv.error = "Selecione uma embarcacao valida"
    vessel_dv.errorTitle = "Embarcacao Invalida"
    ws_viagens.add_data_validation(vessel_dv)
    vessel_dv.add("A3:A1000")

    # Add example data (starting from row 3)
    example_data = [
        ["SURFER 1931", "06:20", "TMIB +16/M9 -16 +13/TMIB (-13) +20/M9 -6/M4 -4/M2 -5/M1 -5"],
        ["SURFER 1871", "06:30", "TMIB +16/M9/M5 -4/PGA3 -4/PGA4 -8"],
        ["AQUA HELIX", "07:00", "TMIB +64/M9 -64 +19/TMIB -19"],
        ["SURFER 1870", "07:30", "TMIB +14/M6 -1/M9/B2 -7/B1 -6"],
        ["SURFER 1930", "07:40", "TMIB +23/M9/M10 -15/M3 -8"],
        ["SURFER 1930", "05:20", "M9 +22/M6 (-11) {TMIB:+1}/B1 (-5)/TMIB (-6) {M6:-1}"],
    ]

    for row_idx, row_data in enumerate(example_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws_viagens.cell(row=row_idx, column=col_idx, value=value)

    wb.save(path)
    print(f"Template criado: {path}")


# =========================
# LEITURA DO EXCEL DE ENTRADA
# =========================
@dataclass
class TripDef:
    vessel: str
    start_hhmm: str
    route: str


def load_trips_from_excel(path: str) -> List[TripDef]:
    """Carrega viagens do arquivo Excel de entrada."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Arquivo de entrada nao encontrado: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb["Viagens"]

    trips: List[TripDef] = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        vessel = str(row[0]).strip()
        departure = row[1]
        route = str(row[2]).strip() if row[2] else ""

        if not route:
            continue

        # Handle time format (might be datetime object)
        if hasattr(departure, 'strftime'):
            departure = departure.strftime("%H:%M")
        elif departure is None:
            departure = "00:00"
        elif ":" not in str(departure):
            # Try to parse as decimal hours (Excel stores times as fractions)
            try:
                hours = float(departure) * 24
                h = int(hours)
                m = int((hours - h) * 60)
                departure = f"{h:02d}:{m:02d}"
            except:
                departure = "00:00"
        else:
            departure = str(departure).strip()

        trips.append(TripDef(vessel=vessel, start_hhmm=departure, route=route))

    return trips


# =========================
# EXCEL OUTPUT
# =========================
def apply_layout(ws):
    # Disable gridlines
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="4F4F4F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(bottom=thin)
    bottom_right_border = Border(bottom=thin, right=thin)
    bottom_left_border = Border(bottom=thin, left=thin)

    # Row 6: Title merged B6:L6
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=12)
    c = ws.cell(row=6, column=2, value="PROGRAMAÇÃO DISTRIBUIÇÃO LOGÍSTICA - TMCP")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 9: Bacia, Data fields
    ws.cell(row=9, column=2, value="Bacia:")
    ws.merge_cells(start_row=9, start_column=3, end_row=9, end_column=5)
    sergipe_cell = ws.cell(row=9, column=3, value="SERGIPE")
    sergipe_cell.border = bottom_border
    ws.cell(row=9, column=4).border = bottom_border
    ws.cell(row=9, column=5).border = bottom_border

    data_cell = ws.cell(row=9, column=7, value="Data:")
    data_cell.font = Font(bold=True)

    ws.cell(row=9, column=8).border = bottom_right_border
    ws.cell(row=9, column=9).border = bottom_border
    ws.cell(row=9, column=10).border = bottom_left_border

    # Row 11: CL Resp.
    ws.cell(row=11, column=2, value="CL Resp.:")
    ws.merge_cells(start_row=11, start_column=3, end_row=11, end_column=5)
    ws.cell(row=11, column=3).border = bottom_border
    ws.cell(row=11, column=4).border = bottom_border
    ws.cell(row=11, column=5).border = bottom_border

    # Row 12: Warning text with "operações com mergulho" in bold/red
    ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=12)

    # Create rich text with "operações com mergulho" in bold/red
    normal_font = InlineFont(sz=11)
    bold_red_font = InlineFont(b=True, color="FF0000", sz=11)

    rich_text = CellRichText(
        TextBlock(normal_font, "Todas as aproximações a unidades marítimas, ainda que desabitadas, devem ser previamente "
                  "informadas no canal 16 do VHF, de modo a verificar se há algum impedimento não devidamente "
                  "informado ou com comunicação intermitente, a exemplo de "),
        TextBlock(bold_red_font, "operações com mergulho"),
        TextBlock(normal_font, " com embarcações não tripuladas. Essa medida visa evitar acidentes por falta de comunicação.")
    )

    warning_cell = ws.cell(row=12, column=2, value=rich_text)
    warning_cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Row 13: PROGRAMAÇÃO DE DISTRIBUIÇÃO DE PAX with light blue background
    light_blue_fill = PatternFill("solid", fgColor="BDD7EE")
    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=12)
    pax_title_cell = ws.cell(row=13, column=2, value="PROGRAMAÇÃO DE DISTRIBUIÇÃO DE PAX")
    pax_title_cell.font = Font(bold=True, size=12)
    pax_title_cell.fill = light_blue_fill
    pax_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 13):
        ws.cell(row=13, column=col).fill = light_blue_fill

    # Row 15: Column headers
    header_row = 15
    fill = PatternFill("solid", fgColor="BDD7EE")
    for col, name in enumerate(COLS, start=2):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[6].height = 24.0
    ws.row_dimensions[12].height = 48.0
    ws.row_dimensions[13].height = 24.0
    ws.row_dimensions[15].height = 57.75


def write_trip_block(ws, start_row: int, vessel: str, summary: str, rows: List[Dict]) -> int:
    thin = Side(style="thin", color="4F4F4F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = PatternFill("solid", fgColor="C6E0B4")
    vessel_cell_fill = PatternFill("solid", fgColor="9BBB59")
    light_pink_fill = PatternFill("solid", fgColor="FFDDDD")

    r = start_row

    # Vessel cell in column B (2)
    ws.cell(row=r, column=2, value=vessel).fill = vessel_cell_fill
    ws.cell(row=r, column=2).font = Font(bold=True, color="9C0006")
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=2).border = border

    # Summary merged from C to L (columns 3-12)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
    ws.cell(row=r, column=3, value=f"{summary}").fill = green
    ws.cell(row=r, column=3).font = Font(bold=True)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=r, column=3).border = border

    for col in range(3, 13):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).fill = green

    ws.row_dimensions[r].height = 18
    r += 1

    is_first_row = True
    for row in rows:
        # Values starting from column B (2)
        # Column mapping: B=EMBARCACAO(empty), C=HR SAIDA, D=QUANT PAX EMB, E=ORIGEM,
        # F=HORA CHEGADA, G=DESTINO, H=QUANT PAX DESEMB, I=QUANT A BORDO, J=LINGADAS,
        # K=OP TRANSBORDO, L=OPERACAO
        values = [
            "",  # EMBARCACAO (empty for data rows)
            row["HR SAIDA ORIGEM"],
            row["QUANT PAX EMBARQUE"],
            row["ORIGEM"],
            row["HORA CHEGADA DESTINO"],
            row["DESTINO"],
            row["QUANT. PAX DESEMBARQUE"],
            row["QUANT PAX A BORDO"],
            "",  # LINGADAS
            row["OPERACAO DE TRANSBORDO"],
            row["OPERACAO"],
        ]
        for col, v in enumerate(values, start=2):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = border

            # Time of departure (column C = col 3) in red/bold ONLY for first row
            if col == 3:
                if is_first_row:
                    cell.font = Font(bold=True, color="FF0000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # QUANT PAX EMBARQUE (column D = col 4) in red/bold with light pink background
            elif col == 4:
                cell.font = Font(bold=True, color="FF0000")
                cell.fill = light_pink_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # QUANT PAX DESEMBARQUE (column H = col 8) with light pink background
            elif col == 8:
                cell.fill = light_pink_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col == 6:  # HORA CHEGADA DESTINO
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 7:  # DESTINO (platform names) - center
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col in (9, 11):  # QUANT A BORDO, OP TRANSBORDO
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 18
        r += 1
        is_first_row = False

    r += 1
    return r


def set_column_widths(ws):
    """Set specific column widths."""
    widths = {
        'A': 11.85546875,
        'B': 16.0,
        'C': 16.7109375,
        'D': 11.28515625,
        'E': 8.42578125,
        'F': 14.7109375,
        'G': 9.28515625,
        'H': 15.28515625,
        'I': 12.7109375,
        'J': 10.0,
        'K': 13.0,
        'L': 25.0,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def safe_save(wb: Workbook, filename: str) -> str:
    try:
        wb.save(filename)
        return filename
    except PermissionError:
        base, ext = os.path.splitext(filename)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}_{ts}{ext}"
        wb.save(alt)
        return alt


def write_extra_vessels_and_observations(ws, start_row: int) -> int:
    """Write the three extra vessel lines and observation section."""
    thin = Side(style="thin", color="4F4F4F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    light_blue_fill = PatternFill("solid", fgColor="BDD7EE")

    r = start_row

    # Three formatted lines with vessels in red/bold, merge C to L
    extra_vessels = ["BARU TAURUS", "C-ITACURUÇÁ", "COSTA OCEÂNICA"]
    for vessel in extra_vessels:
        # Vessel cell in column B
        vessel_cell = ws.cell(row=r, column=2, value=vessel)
        vessel_cell.border = border
        vessel_cell.font = Font(bold=True, color="FF0000")

        # Merge C to L (columns 3-12)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
        for col in range(3, 13):
            ws.cell(row=r, column=col).border = border

        ws.row_dimensions[r].height = 18
        r += 1

    # Skip one line
    r += 1

    # Merge B to L and write OBSERVAÇÃO centered with light blue background
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    obs_cell = ws.cell(row=r, column=2, value="OBSERVAÇÃO")
    obs_cell.font = Font(bold=True)
    obs_cell.fill = light_blue_fill
    obs_cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 13):
        ws.cell(row=r, column=col).fill = light_blue_fill
    r += 1

    # Observation lines - merge B to L for each
    observations = [
        "*Não iniciar a navegação com pax que não esteja no manifesto sem a autorização do CL",
        "*Não coletar carga que não esteja no manifesto sem a autorização do CL;",
        "*Essa programação pode sofrer alterações de horário, origem, destino, roteiro e embarcação.",
        "*Caso haja alguma outra demanda emergencial, deve ser enviado um e-mail pelo GIO da plataforma, para a célula do Planejamento PTMA2 (cc-ptma2-prog.transp.se-ba@petrobras.com.br) e para APMA9 (Controlador Logístico de PCM-09)",
    ]

    for i, obs in enumerate(observations):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
        obs_line_cell = ws.cell(row=r, column=2, value=obs)
        obs_line_cell.alignment = Alignment(wrap_text=True, vertical="center")
        # Last observation has height 28.5, others have 18.0
        if i == len(observations) - 1:
            ws.row_dimensions[r].height = 28.5
        else:
            ws.row_dimensions[r].height = 18.0
        r += 1

    return r


# =========================
# MAIN
# =========================
def main():
    dist = load_distances_json(DIST_FILE)
    speeds = load_speeds(SPEED_FILE)

    # Criar template se nao existir
    if not os.path.exists(INPUT_FILE):
        print(f"Arquivo de entrada '{INPUT_FILE}' nao encontrado.")
        print("Criando template...")
        create_input_template(INPUT_FILE, speeds)
        print(f"\nPreencha o arquivo '{INPUT_FILE}' e execute novamente.")
        return

    trips = load_trips_from_excel(INPUT_FILE)

    if not trips:
        print("Nenhuma viagem encontrada no arquivo de entrada.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    apply_layout(ws)

    row_ptr = 16  # Start data at row 16 (after header row 15)

    for trip in trips:
        vessel_upper = trip.vessel.upper()
        speed_kn = None

        variations = [
            vessel_upper,
            vessel_upper.replace("_", " "),
            vessel_upper.replace("_", ""),
            vessel_upper.replace(" ", "_"),
        ]
        for var in variations:
            if var in speeds:
                speed_kn = speeds[var]
                break

        if speed_kn is None:
            speed_kn = DEFAULT_SPEED_KN
            print(f"AVISO: Velocidade nao encontrada para '{trip.vessel}'. Usando default: {DEFAULT_SPEED_KN} nos")

        # Parse route string to stops
        stops = parse_route(trip.route)

        rows, total, summary, vessel_speed = simulate_trip(
            dist=dist,
            vessel_name=trip.vessel,
            start_hhmm=trip.start_hhmm,
            stops=stops,
            speed_kn=speed_kn,
            minutes_per_pax=MINUTES_PER_PAX,
        )

        summary_compact = summary.replace(" > ", ">")
        row_ptr = write_trip_block(
            ws=ws,
            start_row=row_ptr,
            vessel=trip.vessel,
            summary=summary_compact,
            rows=rows,
        )

    # Add extra vessels and observations at the end
    row_ptr = write_extra_vessels_and_observations(ws, row_ptr)

    # Set specific column widths
    set_column_widths(ws)

    saved_as = safe_save(wb, OUT_FILE)
    print(f"\nPlanilha gerada: {saved_as}")


if __name__ == "__main__":
    main()
